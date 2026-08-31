import json
import csv
import hashlib
import io
import shutil
import subprocess
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote
from uuid import uuid4

from fastapi import Body, Depends, FastAPI, File, Form, HTTPException, Request, Response, UploadFile, status
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, func, inspect, or_, select
from sqlalchemy.orm import Session
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.middleware.httpsredirect import HTTPSRedirectMiddleware

from app.core.config import get_settings
from app.core.time import as_china_time, china_now, normalize_json_timestamps
from app.core.permissions import CAPABILITIES, get_current_user, has_capability, may_edit, require_roles, user_capabilities
from app.core.security import create_access_token, hash_password, new_csrf_token, new_totp_secret, require_csrf, verify_password, verify_totp_code
from app.db import SessionLocal, engine, get_db, init_db
from app.models import AiConversation, AiConversationMessage, AiEvaluationRun, AiPendingAction, AuditLog, AuditReversal, BackgroundTask, CandidateStatus, DeletedStudent, ExportTemplate, FieldProvenance, HighRiskApproval, ImportBatch, ImportMappingTemplate, ImportMatchReview, ImportPreview, LoginSecurityEvent, QualityIssueCase, QualityScan, RelatedInfoCandidate, Role, SavedStudentFilter, SourceDocument, Student, StudentMerge, StudentRelatedInfoCard, StudentVersion, SystemAlert, SystemBackup, TemplateRevision, User, UserDataScope, UserDataScopeRule, WordImportCandidate, utcnow
from app.schemas import AdministratorCreate, AdministratorUpdate, AiQuestion, CandidateApproval, DataScopeUpdate, ExcelImportCommit, ExportTemplateInput, ImportTemplateInput, LoginRequest, MfaCode, PasswordChange, StudentCreate, StudentDeletion, StudentResponse, StudentUpdate, SystemControlsUpdate, SystemSettingsUpdate
from app.services.ai import AGGREGATE_FIELD_LABELS, RESPONSE_FIELD_LABELS, _is_related_info_query, express_assistant_answer, get_ai_health, plan_assistant_question
from app.services.audit import audit, verify_audit_chain
from app.services.backups import create_database_backup, delete_database_backup, drill_restore_backup, maybe_create_scheduled_backup, restore_backup, validate_backup
from app.services.exports import EXPORT_FIELD_HEADERS, SENSITIVE_EXPORT_FIELDS, create_student_export
from app.services.files import document_path, register_upload
from app.services.source_downloads import create_student_scoped_source_copy, sanitized_download_filename
from app.services.imports import apply_related_info_candidate, apply_word_candidate, dismiss_import_match_review, import_excel, import_related_info, import_word_for_review, preview_excel_import, resolve_import_match_review
from app.services.ai_evaluations import run_ai_regression, serialize_ai_evaluation
from app.services.monitoring import evaluate_alerts, serialize_alert, system_snapshot
from app.services.students import DATE_FILTER_FIELDS, STUDENT_FIELDS, build_student_query, create_student, get_provenance, list_deleted_students, list_student_filter_options, list_student_versions, list_students, list_students_page, permanently_delete_student, purge_expired_deleted_students, record_student_version, restore_deleted_student, restore_student_version, student_to_dict, update_student
from app.services.preferences import get_controls, set_controls
from app.services.quality import run_quality_scan, serialize_quality_scan
from app.services.scopes import apply_scope, ensure_new_student_scope, ensure_student_scope, get_effective_user_scopes, get_user_scope, get_user_scopes, scope_as_dict, serialize_student_for_user
from app.services.tasks import serialize_task, submit_task, update_task
from app.services.governance import list_duplicate_groups, list_student_reminders, merge_students, rollback_import_batch, rollback_related_info_batch, student_timeline
from app.services.updates import check_for_update, create_update_job, get_update_configuration, get_update_status, launch_update_runner, normalize_repository, save_update_configuration, update_run_root, write_update_status
from app.version import APP_RELEASE


class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Referrer-Policy"] = "same-origin"
        response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
        response.headers["Content-Security-Policy"] = "default-src 'self'; script-src 'self'; style-src 'self'; img-src 'self' data:; connect-src 'self'; font-src 'self'"
        return response


class ChinaTimeResponseMiddleware(BaseHTTPMiddleware):
    """Attach China Standard Time offsets to every JSON timestamp sent to clients."""

    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        content_type = str(response.headers.get("content-type") or "")
        if "application/json" not in content_type or any(key.lower() == b"set-cookie" for key, _ in response.raw_headers):
            return response
        body = b"".join([chunk async for chunk in response.body_iterator])
        try:
            payload = json.loads(body)
            normalized = normalize_json_timestamps(payload)
            body = json.dumps(normalized, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
        except (TypeError, ValueError, json.JSONDecodeError):
            pass
        headers = {key: value for key, value in response.headers.items() if key.lower() != "content-length"}
        return Response(content=body, status_code=response.status_code, headers=headers, media_type="application/json", background=response.background)


settings = get_settings()
app = FastAPI(
    title=settings.app_name,
    docs_url=None if settings.is_production else "/docs",
    redoc_url=None,
    json_encoders={datetime: lambda value: as_china_time(value).isoformat() if as_china_time(value) else None},
)
app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(ChinaTimeResponseMiddleware)
if settings.is_production:
    app.add_middleware(HTTPSRedirectMiddleware)
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")


@app.on_event("startup")
def startup() -> None:
    settings.storage_path.mkdir(parents=True, exist_ok=True)
    settings.export_path.mkdir(parents=True, exist_ok=True)
    settings.backup_path.mkdir(parents=True, exist_ok=True)
    init_db()
    db = SessionLocal()
    try:
        maybe_create_scheduled_backup(db)
        purge_expired_deleted_students(db)
        evaluate_alerts(db)
        db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()


def _serialize_student(student: Student) -> dict[str, Any]:
    return StudentResponse.model_validate(student).model_dump(mode="json")


def _student_scope(db: Session, user: User) -> list[dict[str, str]]:
    return get_effective_user_scopes(db, user) if user.role != Role.SUPER_ADMIN else []


def _count_students(db: Session, filters: dict[str, Any] | None = None, scope: list[dict[str, str]] | None = None) -> int:
    """Count a filtered student query without losing its FROM clause."""
    students = build_student_query(filters=filters or {}, scope=scope).order_by(None).subquery()
    return int(db.scalar(select(func.count()).select_from(students)) or 0)


def _serialize_student_for_user(db: Session, user: User, student: Student) -> dict[str, Any]:
    return serialize_student_for_user(db, user, _serialize_student(student))


def _set_session_cookies(response: Response, user: User) -> None:
    access_token = create_access_token(user.id, user.role.value, user.session_version)
    max_age = settings.jwt_expire_minutes * 60
    cookie_options = {"secure": settings.cookie_secure, "samesite": "lax", "path": "/", "max_age": max_age}
    response.set_cookie("access_token", access_token, httponly=True, **cookie_options)
    response.set_cookie("csrf_token", new_csrf_token(), httponly=False, **cookie_options)
    # This is not an authorization token. It lets the browser start a distinct
    # AI conversation for every successful login while keeping refreshes intact.
    response.set_cookie("ai_session_id", uuid4().hex, httponly=False, **cookie_options)


def _require_edit_access(user: User) -> None:
    if not may_edit(user):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="没有编辑权限")


def _require_capability(user: User, capability: str, label: str) -> None:
    if not has_capability(user, capability):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail=f"没有{label}权限")


def _admin_can_view_teacher_record(db: Session, user: User, actor_id: int | None) -> bool:
    if user.role == Role.SUPER_ADMIN:
        return True
    actor = db.get(User, actor_id) if actor_id else None
    return bool(actor and actor.role == Role.TEACHER)


def _teacher_can_access_import(user: User, imported_by_id: int | None) -> bool:
    """Teachers may only read or review records created by their own imports."""
    return user.role != Role.TEACHER or imported_by_id == user.id


def _require_teacher_import_access(user: User, imported_by_id: int | None) -> None:
    if not _teacher_can_access_import(user, imported_by_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="教师只能查看和审核自己导入的相关信息")


def _scope_student_ids(db: Session, user: User, student_ids: set[int]) -> set[int]:
    """Apply the account data range before returning related-info records."""
    if not student_ids or user.role == Role.SUPER_ADMIN:
        return student_ids
    statement = apply_scope(select(Student.id).where(Student.id.in_(student_ids)), db, user)
    return {int(student_id) for student_id in db.scalars(statement)}


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    response = templates.TemplateResponse(request=request, name="login.html")
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response


@app.get("/", response_class=HTMLResponse)
def app_page(request: Request):
    response = templates.TemplateResponse(request=request, name="index.html")
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response


def _request_ip(request: Request) -> str | None:
    """Use the proxy header only for deployed instances, never blindly in local mode."""
    if settings.is_production:
        forwarded = str(request.headers.get("x-forwarded-for") or "").split(",", 1)[0].strip()
        if forwarded:
            return forwarded[:64]
    return request.client.host[:64] if request.client else None


def _network_key(ip_address: str | None) -> str | None:
    if not ip_address:
        return None
    if "." in ip_address:
        parts = ip_address.split(".")
        return ".".join(parts[:3]) + ".0/24" if len(parts) == 4 else ip_address
    if ":" in ip_address:
        return ":".join(ip_address.split(":")[:4]) + "::/64"
    return ip_address


def _device_label(request: Request) -> str:
    agent = str(request.headers.get("user-agent") or "")
    platform = "Windows" if "Windows" in agent else "macOS" if "Mac OS" in agent else "Android" if "Android" in agent else "iPhone/iPad" if any(token in agent for token in ("iPhone", "iPad")) else "未知设备"
    browser = "Edge" if "Edg/" in agent else "Chrome" if "Chrome/" in agent else "Firefox" if "Firefox/" in agent else "Safari" if "Safari/" in agent else "浏览器"
    return f"{platform} · {browser}"


def _record_login_security_event(
    db: Session,
    request: Request,
    username: str,
    event_type: str,
    user: User | None = None,
    unusual: bool = False,
) -> LoginSecurityEvent:
    ip_address = _request_ip(request)
    event = LoginSecurityEvent(
        user_id=user.id if user else None,
        username=str(username or "")[:64],
        event_type=event_type,
        ip_address=ip_address,
        network_key=_network_key(ip_address),
        device_label=_device_label(request),
        user_agent=str(request.headers.get("user-agent") or "")[:512] or None,
        is_unusual=unusual,
    )
    db.add(event)
    return event


@app.post("/api/auth/login")
def login(payload: LoginRequest, request: Request, response: Response, db: Session = Depends(get_db)) -> dict[str, Any]:
    user = db.scalar(select(User).where(User.username == payload.username))
    now = utcnow()
    if not user or not user.is_active:
        _record_login_security_event(db, request, payload.username, "login_failed")
        audit(db, "login_failed", "user", payload.username, request=request)
        db.commit()
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="用户名或密码错误")
    locked_until = user.locked_until
    locked_until = as_china_time(locked_until)
    if locked_until and locked_until > now:
        _record_login_security_event(db, request, user.username, "login_locked", user)
        db.commit()
        raise HTTPException(status_code=status.HTTP_423_LOCKED, detail="登录尝试过多，请稍后再试")
    if not verify_password(payload.password, user.password_hash):
        user.failed_login_count += 1
        if user.failed_login_count >= 5:
            user.locked_until = now + timedelta(minutes=15)
            user.failed_login_count = 0
        _record_login_security_event(db, request, user.username, "login_failed", user)
        audit(db, "login_failed", "user", user.id, actor=user, request=request)
        db.commit()
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="用户名或密码错误")
    if user.mfa_enabled and not verify_totp_code(user.mfa_secret, payload.mfa_code):
        _record_login_security_event(db, request, user.username, "login_mfa_failed", user)
        audit(db, "login_mfa_failed", "user", user.id, actor=user, request=request)
        db.commit()
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="二次验证代码不正确")
    user.failed_login_count = 0
    user.locked_until = None
    user.last_login_at = now
    current_network = _network_key(_request_ip(request))
    recent_networks = set(
        db.scalars(
            select(LoginSecurityEvent.network_key).where(
                LoginSecurityEvent.user_id == user.id,
                LoginSecurityEvent.event_type == "login_success",
                LoginSecurityEvent.created_at >= now - timedelta(days=90),
                LoginSecurityEvent.network_key.is_not(None),
            )
        )
    )
    unusual = bool(recent_networks and current_network and current_network not in recent_networks)
    _record_login_security_event(db, request, user.username, "login_success", user, unusual=unusual)
    audit(db, "unusual_login" if unusual else "login", "user", user.id, actor=user, after={"network_changed": unusual}, request=request)
    db.commit()
    _set_session_cookies(response, user)
    security_notice = "检测到该账号从新的网络位置登录，请确认是本人操作。" if unusual else None
    return {"user": {"id": user.id, "username": user.username, "display_name": user.display_name, "role": user.role.value, "capabilities": user_capabilities(user), "must_change_password": user.must_change_password, "mfa_enabled": user.mfa_enabled}, "security_notice": security_notice}


@app.post("/api/auth/logout")
def logout(request: Request, response: Response, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, bool]:
    require_csrf(request)
    audit(db, "logout", "user", user.id, actor=user, request=request)
    db.commit()
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("csrf_token", path="/")
    response.delete_cookie("ai_session_id", path="/")
    return {"ok": True}


@app.put("/api/auth/password")
def change_own_password(
    payload: PasswordChange,
    request: Request,
    response: Response,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    if not verify_password(payload.current_password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前密码不正确")
    if payload.new_password != payload.confirm_password:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="两次输入的新密码不一致")
    user.password_hash = hash_password(payload.new_password)
    user.password_changed_at = utcnow()
    user.must_change_password = False
    user.failed_login_count = 0
    user.locked_until = None
    user.session_version = max(1, int(user.session_version or 1)) + 1
    audit(db, "change_own_password", "user", user.id, actor=user, after={"password_changed": True}, request=request)
    db.commit()
    _set_session_cookies(response, user)
    return {"ok": True, "message": "密码已更新，其他已登录会话已失效", "must_change_password": False}


@app.post("/api/auth/mfa/setup")
def setup_mfa(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, str]:
    require_csrf(request)
    if user.mfa_enabled:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="二次验证已经启用")
    user.mfa_secret = new_totp_secret()
    audit(db, "setup_mfa", "user", user.id, actor=user, after={"prepared": True}, request=request)
    db.commit()
    return {"secret": user.mfa_secret, "issuer": settings.app_name, "account": user.username}


@app.post("/api/auth/mfa/enable")
def enable_mfa(
    payload: MfaCode,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, bool]:
    require_csrf(request)
    if not user.mfa_secret:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="请先生成身份验证器密钥")
    if not verify_totp_code(user.mfa_secret, payload.code):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="验证代码不正确")
    user.mfa_enabled = True
    audit(db, "enable_mfa", "user", user.id, actor=user, after={"enabled": True}, request=request)
    db.commit()
    return {"ok": True}


@app.post("/api/auth/mfa/disable")
def disable_mfa(
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, bool]:
    require_csrf(request)
    if not verify_password(str(payload.get("current_password") or ""), user.password_hash):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前密码不正确")
    if not verify_totp_code(user.mfa_secret, str(payload.get("code") or "")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="二次验证代码不正确")
    user.mfa_secret = None
    user.mfa_enabled = False
    audit(db, "disable_mfa", "user", user.id, actor=user, after={"enabled": False}, request=request)
    db.commit()
    return {"ok": True}


@app.get("/api/me")
def current_user(user: User = Depends(get_current_user)) -> dict[str, Any]:
    return {"id": user.id, "username": user.username, "display_name": user.display_name, "role": user.role.value, "capabilities": user_capabilities(user), "must_change_password": user.must_change_password, "mfa_enabled": user.mfa_enabled}


@app.get("/api/system/settings")
def get_system_settings(user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN))) -> dict[str, str]:
    return {"username": user.username, "display_name": user.display_name, "role": user.role.value}


def _serialize_administrator(administrator: User) -> dict[str, Any]:
    return {
        "id": administrator.id,
        "username": administrator.username,
        "display_name": administrator.display_name,
        "role": administrator.role.value,
        "is_active": administrator.is_active,
        "created_at": administrator.created_at,
        "last_login_at": administrator.last_login_at,
        "permissions": administrator.permissions,
        "capabilities": user_capabilities(administrator),
        "must_change_password": administrator.must_change_password,
        "mfa_enabled": administrator.mfa_enabled,
    }


def _normalize_permissions(values: list[str] | None) -> list[str] | None:
    if values is None:
        return None
    normalized = sorted({str(value).strip() for value in values if str(value).strip()})
    invalid = [value for value in normalized if value not in CAPABILITIES]
    if invalid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"存在未知权限：{'、'.join(invalid)}")
    return normalized


@app.get("/api/system/administrators")
def list_administrators(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> list[dict[str, Any]]:
    statement = select(User).where(User.role.in_([Role.ADMIN, Role.TEACHER])) if user.role == Role.SUPER_ADMIN else select(User).where(User.role == Role.TEACHER)
    administrators = list(db.scalars(statement.order_by(User.role.asc(), User.created_at.desc(), User.id.desc())))
    return [_serialize_administrator(administrator) for administrator in administrators]


@app.post("/api/system/administrators", status_code=status.HTTP_201_CREATED)
def create_administrator(
    payload: AdministratorCreate,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    if not verify_password(payload.current_password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前管理员密码不正确")
    if payload.password != payload.confirm_password:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="两次输入的管理员密码不一致")
    if payload.role == Role.SUPER_ADMIN.value or (user.role == Role.ADMIN and payload.role != Role.TEACHER.value):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="普通管理员只能管理教师账号")
    if db.scalar(select(User).where(User.username == payload.username)):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="用户名已被使用")

    administrator = User(
        username=payload.username,
        display_name=payload.display_name,
        password_hash=hash_password(payload.password),
        role=Role(payload.role),
        permissions=_normalize_permissions(payload.permissions),
        is_active=True,
        must_change_password=True,
        password_changed_at=utcnow(),
    )
    db.add(administrator)
    db.flush()
    audit(
        db,
        "create_administrator",
        "user",
        administrator.id,
        actor=user,
        after={"username": administrator.username, "display_name": administrator.display_name, "role": administrator.role.value, "permissions": administrator.permissions},
        request=request,
    )
    db.commit()
    return _serialize_administrator(administrator)


@app.put("/api/system/administrators/{administrator_id}")
def update_administrator(
    administrator_id: int,
    payload: AdministratorUpdate,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    administrator = db.get(User, administrator_id)
    if not administrator or administrator.role == Role.SUPER_ADMIN or (user.role == Role.ADMIN and administrator.role != Role.TEACHER):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="账号不存在")
    if not verify_password(payload.current_password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前管理员密码不正确")
    if payload.new_password != payload.confirm_password:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="两次输入的新密码不一致")
    username_owner = db.scalar(select(User).where(User.username == payload.username, User.id != administrator.id))
    if username_owner:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="用户名已被使用")
    if payload.role == Role.SUPER_ADMIN.value or (user.role == Role.ADMIN and payload.role != Role.TEACHER.value):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="普通管理员只能管理教师账号")

    before = {"username": administrator.username, "display_name": administrator.display_name, "role": administrator.role.value, "permissions": administrator.permissions}
    administrator.username = payload.username
    administrator.display_name = payload.display_name
    administrator.role = Role(payload.role)
    administrator.permissions = _normalize_permissions(payload.permissions)
    if payload.new_password:
        administrator.password_hash = hash_password(payload.new_password)
        administrator.failed_login_count = 0
        administrator.locked_until = None
        administrator.must_change_password = True
        administrator.password_changed_at = utcnow()
        administrator.session_version = max(1, int(administrator.session_version or 1)) + 1
    audit(
        db,
        "update_administrator",
        "user",
        administrator.id,
        actor=user,
        before=before,
        after={"username": administrator.username, "display_name": administrator.display_name, "role": administrator.role.value, "permissions": administrator.permissions, "password_reset": bool(payload.new_password)},
        request=request,
    )
    db.commit()
    return _serialize_administrator(administrator)


@app.put("/api/system/settings")
def update_system_settings(
    payload: SystemSettingsUpdate,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, str]:
    require_csrf(request)
    if not verify_password(payload.current_password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前密码不正确")
    if payload.new_password != payload.confirm_password:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="两次输入的新密码不一致")
    username_owner = db.scalar(select(User).where(User.username == payload.username, User.id != user.id))
    if username_owner:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="用户名已被使用")
    before = {"username": user.username, "display_name": user.display_name}
    user.username = payload.username
    user.display_name = payload.display_name
    if payload.new_password:
        user.password_hash = hash_password(payload.new_password)
        user.failed_login_count = 0
        user.locked_until = None
        user.password_changed_at = utcnow()
        user.must_change_password = False
    audit(db, "update_system_settings", "user", user.id, actor=user, before=before, after={"username": user.username, "display_name": user.display_name, "password_changed": bool(payload.new_password)}, request=request)
    db.commit()
    return {"username": user.username, "display_name": user.display_name}


@app.post("/api/system/high-risk/authorize")
def authorize_high_risk_setting(
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    action = str(payload.get("action") or "").strip()
    if action != "clear_all_students":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="未知的高危操作")
    username = str(payload.get("username") or "").strip()
    password = str(payload.get("password") or "")
    approver = db.scalar(select(User).where(User.username == username, User.role == Role.SUPER_ADMIN, User.is_active.is_(True)))
    if not approver or not verify_password(password, approver.password_hash):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="超级管理员账号或密码不正确")
    if approver.locked_until:
        locked_until = as_china_time(approver.locked_until)
        if locked_until > utcnow():
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="该超级管理员账号当前已锁定")
    db.execute(delete(HighRiskApproval).where(HighRiskApproval.requested_by_id == user.id, HighRiskApproval.expires_at < utcnow()))
    approval = HighRiskApproval(
        id=str(uuid4()),
        requested_by_id=user.id,
        approved_by_id=approver.id,
        action=action,
        expires_at=utcnow() + timedelta(minutes=5),
    )
    db.add(approval)
    audit(
        db,
        "authorize_high_risk_setting",
        "high_risk_approval",
        approval.id,
        actor=user,
        after={"action": action, "approved_by_id": approver.id, "expires_at": approval.expires_at.isoformat()},
        request=request,
    )
    db.commit()
    return {"approval_id": approval.id, "action": approval.action, "expires_at": approval.expires_at}


@app.post("/api/system/high-risk/clear-students")
def clear_all_students_high_risk(
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    approval_id = str(payload.get("approval_id") or "").strip()
    approval = db.get(HighRiskApproval, approval_id)
    if not approval or approval.requested_by_id != user.id or approval.action != "clear_all_students":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="高危操作授权无效，请重新验证超级管理员凭证")
    approval_expiry = as_china_time(approval.expires_at)
    if approval.used_at or approval_expiry <= utcnow():
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="高危操作授权已过期或已经使用，请重新验证")
    if int(payload.get("confirmation_count") or 0) != 3 or str(payload.get("confirmation_phrase") or "").strip() != "永久清空学生档案":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="需要完成三次确认，并输入“永久清空学生档案”")
    students = list(db.scalars(select(Student).order_by(Student.id.asc())))
    student_refs = [{"id": student.id, "student_no": student.student_no, "full_name": student.full_name} for student in students]
    recycle_ids: list[int] = []
    for student in students:
        recycle_ids.append(permanently_delete_student(db, student, user).id)
    approval.used_at = utcnow()
    audit(
        db,
        "clear_all_students_high_risk",
        "student_archive",
        "all",
        actor=user,
        after={
            "deleted_student_count": len(student_refs),
            "students": student_refs[:200],
            "recycle_ids": recycle_ids,
            "approval_id": approval.id,
            "approved_by_id": approval.approved_by_id,
            "confirmation_count": 3,
        },
        request=request,
    )
    db.commit()
    return {"ok": True, "deleted_students": len(student_refs), "recycle_bin": True}


@app.post("/api/system/users/{account_id}/revoke-sessions")
def revoke_user_sessions(
    account_id: int,
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, bool]:
    require_csrf(request)
    if str(payload.get("confirmation_phrase") or "").strip() != "注销会话":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请输入“注销会话”确认")
    if not verify_password(str(payload.get("current_password") or ""), user.password_hash):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前管理员密码不正确")
    account = db.get(User, account_id)
    if not account or account.role == Role.SUPER_ADMIN or (user.role == Role.ADMIN and account.role != Role.TEACHER):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="账号不存在")
    account.session_version = max(1, int(account.session_version or 1)) + 1
    audit(db, "revoke_user_sessions", "user", account.id, actor=user, after={"session_version": account.session_version}, request=request)
    db.commit()
    return {"ok": True}


@app.get("/api/system/controls")
def system_controls(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, bool]:
    return get_controls(db)


@app.put("/api/system/controls")
def update_system_controls(
    payload: SystemControlsUpdate,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, bool]:
    require_csrf(request)
    before = get_controls(db)
    controls = set_controls(db, payload.model_dump(), user)
    audit(db, "update_system_controls", "system_preference", "security_controls", actor=user, before=before, after=controls, request=request)
    db.commit()
    return controls


def _serialize_scope(scope: UserDataScope | None) -> dict[str, str]:
    return scope_as_dict(scope)


def _scope_mode(db: Session, account: User) -> str:
    if get_user_scopes(db, account):
        return "restricted"
    return "all"


@app.get("/api/system/data-scopes")
def list_data_scopes(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> list[dict[str, Any]]:
    statement = select(User).where(User.role.in_([Role.ADMIN, Role.TEACHER])) if user.role == Role.SUPER_ADMIN else select(User).where(User.role == Role.TEACHER)
    accounts = list(db.scalars(statement.order_by(User.role.asc(), User.username.asc())))
    return [{"id": account.id, "username": account.username, "display_name": account.display_name, "role": account.role.value, "is_active": account.is_active, "scope": get_user_scopes(db, account), "scope_count": len(get_user_scopes(db, account)), "scope_mode": _scope_mode(db, account)} for account in accounts]


@app.put("/api/system/users/{account_id}/data-scope")
def update_data_scope(
    account_id: int,
    payload: DataScopeUpdate,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    account = db.get(User, account_id)
    if not account or account.role == Role.SUPER_ADMIN or (user.role == Role.ADMIN and account.role != Role.TEACHER):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="可配置数据范围的账号不存在")
    raw_rules = [payload.model_dump(exclude={"rules"}), *(item.model_dump() for item in payload.rules)]
    rules = [{key: str(value).strip() if value else None for key, value in item.items()} for item in raw_rules]
    rules = [item for item in rules if any(item.values())]
    existing = db.scalar(select(UserDataScope).where(UserDataScope.user_id == account.id))
    before = {"rules": get_user_scopes(db, account), "scope_mode": _scope_mode(db, account)}
    if not rules:
        primary = {key: None for key in ("school", "college", "school_major", "current_class")}
        if existing:
            for key, value in primary.items():
                setattr(existing, key, value)
            existing.updated_by_id = user.id
        else:
            db.add(UserDataScope(user_id=account.id, updated_by_id=user.id, **primary))
    else:
        primary = rules[0]
        if existing:
            for key, value in primary.items():
                setattr(existing, key, value)
            existing.updated_by_id = user.id
        else:
            db.add(UserDataScope(user_id=account.id, updated_by_id=user.id, **primary))
    db.execute(delete(UserDataScopeRule).where(UserDataScopeRule.user_id == account.id))
    for rule in rules[1:]:
        db.add(UserDataScopeRule(user_id=account.id, created_by_id=user.id, **rule))
    after = {"rules": rules, "scope_mode": "all" if not rules else "restricted"}
    audit(db, "update_data_scope", "user", account.id, actor=user, before=before, after=after, request=request)
    db.commit()
    return {"id": account.id, "scope": rules, "scope_mode": after["scope_mode"]}


@app.post("/api/system/users/{account_id}/data-scope/preview")
def preview_data_scope(
    account_id: int,
    payload: DataScopeUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    account = db.get(User, account_id)
    if not account or account.role == Role.SUPER_ADMIN or (user.role == Role.ADMIN and account.role != Role.TEACHER):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="可配置数据范围的账号不存在")
    raw_rules = [payload.model_dump(exclude={"rules"}), *(item.model_dump() for item in payload.rules)]
    rules = [{key: str(value).strip() if value else None for key, value in item.items()} for item in raw_rules]
    rules = [item for item in rules if any(item.values())]
    statement = build_student_query(scope=rules)
    scoped_students = statement.order_by(None).subquery()
    total = int(db.scalar(select(func.count()).select_from(scoped_students)) or 0)
    dimension_counts = {}
    for key, column_name in (("schools", "school"), ("colleges", "college"), ("majors", "school_major"), ("classes", "current_class")):
        column = getattr(scoped_students.c, column_name)
        dimension_counts[key] = int(
            db.scalar(
                select(func.count(func.distinct(column)))
                .select_from(scoped_students)
                .where(column.is_not(None), column != "")
            )
            or 0
        )
    samples = list(db.scalars(statement.limit(5)))
    return {
        "rules": rules,
        "total_students": total,
        "dimension_counts": dimension_counts,
        "samples": [{"student_no": item.student_no, "full_name": item.full_name, "school_major": item.school_major, "current_class": item.current_class} for item in samples],
    }


def _serialize_import_template(template: ImportMappingTemplate) -> dict[str, Any]:
    return {"id": template.id, "name": template.name, "mapping": template.mapping or {}, "required_fields": template.required_fields or [], "default_mode": template.default_mode, "update_policy": template.update_policy, "created_by_id": template.created_by_id, "created_at": template.created_at, "updated_at": template.updated_at}


def _record_template_revision(db: Session, template_kind: str, template_id: int, snapshot: dict[str, Any], action: str, actor: User) -> TemplateRevision:
    latest = db.scalar(
        select(func.max(TemplateRevision.revision_no)).where(
            TemplateRevision.template_kind == template_kind,
            TemplateRevision.template_id == template_id,
        )
    )
    safe_snapshot = json.loads(json.dumps(snapshot, default=lambda value: value.isoformat() if hasattr(value, "isoformat") else str(value)))
    revision = TemplateRevision(
        template_kind=template_kind,
        template_id=template_id,
        revision_no=int(latest or 0) + 1,
        action=action,
        snapshot=safe_snapshot,
        created_by_id=actor.id,
    )
    db.add(revision)
    return revision


def _serialize_template_revision(revision: TemplateRevision, actor_name: str | None = None) -> dict[str, Any]:
    return {
        "id": revision.id,
        "template_kind": revision.template_kind,
        "template_id": revision.template_id,
        "revision_no": revision.revision_no,
        "action": revision.action,
        "snapshot": revision.snapshot or {},
        "created_by": actor_name or "系统",
        "created_at": revision.created_at,
    }


@app.get("/api/template-revisions/{template_kind}/{template_id}")
def list_template_revisions(
    template_kind: str,
    template_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> list[dict[str, Any]]:
    if template_kind not in {"import", "export"}:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="模板类型不存在")
    rows = db.execute(
        select(TemplateRevision, User.display_name, User.username)
        .outerjoin(User, TemplateRevision.created_by_id == User.id)
        .where(TemplateRevision.template_kind == template_kind, TemplateRevision.template_id == template_id)
        .order_by(TemplateRevision.revision_no.desc())
    )
    return [
        _serialize_template_revision(revision, display_name or username)
        for revision, display_name, username in rows
    ]


@app.get("/api/import-templates")
def list_import_templates(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> list[dict[str, Any]]:
    return [_serialize_import_template(item) for item in db.scalars(select(ImportMappingTemplate).order_by(ImportMappingTemplate.name.asc()))]


@app.post("/api/import-templates", status_code=status.HTTP_201_CREATED)
def create_import_template(
    payload: ImportTemplateInput,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    name = payload.name.strip()
    if db.scalar(select(ImportMappingTemplate).where(ImportMappingTemplate.name == name)):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="同名导入模板已存在")
    template = ImportMappingTemplate(name=name, mapping={str(key).upper(): value for key, value in payload.mapping.items() if value}, required_fields=payload.required_fields, default_mode=payload.default_mode, update_policy=payload.update_policy, created_by_id=user.id)
    db.add(template)
    db.flush()
    _record_template_revision(db, "import", template.id, _serialize_import_template(template), "created", user)
    audit(db, "create_import_template", "import_template", template.id, actor=user, after=_serialize_import_template(template), request=request)
    db.commit()
    return _serialize_import_template(template)


@app.put("/api/import-templates/{template_id}")
def update_import_template(
    template_id: int,
    payload: ImportTemplateInput,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    template = db.get(ImportMappingTemplate, template_id)
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入模板不存在")
    owner = db.scalar(select(ImportMappingTemplate).where(ImportMappingTemplate.name == payload.name.strip(), ImportMappingTemplate.id != template.id))
    if owner:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="同名导入模板已存在")
    before = _serialize_import_template(template)
    template.name = payload.name.strip()
    template.mapping = {str(key).upper(): value for key, value in payload.mapping.items() if value}
    template.required_fields = payload.required_fields
    template.default_mode = payload.default_mode
    template.update_policy = payload.update_policy
    db.flush()
    _record_template_revision(db, "import", template.id, _serialize_import_template(template), "updated", user)
    audit(db, "update_import_template", "import_template", template.id, actor=user, before=before, after=_serialize_import_template(template), request=request)
    db.commit()
    return _serialize_import_template(template)


@app.delete("/api/import-templates/{template_id}")
def delete_import_template(
    template_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, bool]:
    require_csrf(request)
    template = db.get(ImportMappingTemplate, template_id)
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入模板不存在")
    snapshot = _serialize_import_template(template)
    _record_template_revision(db, "import", template.id, snapshot, "deleted", user)
    audit(db, "delete_import_template", "import_template", template.id, actor=user, before=snapshot, after={"deleted": True}, request=request)
    db.delete(template)
    db.commit()
    return {"ok": True}


def _serialize_export_template(template: ExportTemplate) -> dict[str, Any]:
    return {"id": template.id, "name": template.name, "fields": template.fields or [], "filters": template.filters or {}, "include_provenance": template.include_provenance, "mask_sensitive": template.mask_sensitive, "created_by_id": template.created_by_id, "created_at": template.created_at, "updated_at": template.updated_at}


def _validate_export_template(payload: ExportTemplateInput) -> tuple[list[str], dict[str, str]]:
    fields = list(dict.fromkeys(field for field in payload.fields if field in EXPORT_FIELD_HEADERS))
    if payload.fields and not fields:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导出模板没有有效字段")
    filters = {str(key): str(value).strip() for key, value in payload.filters.items() if str(key) in EXPORT_FIELD_HEADERS and str(value).strip()}
    return fields, filters


@app.get("/api/export-templates")
def list_export_templates(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[dict[str, Any]]:
    _require_capability(user, "student_export", "导出")
    return [_serialize_export_template(item) for item in db.scalars(select(ExportTemplate).order_by(ExportTemplate.name.asc()))]


@app.post("/api/export-templates", status_code=status.HTTP_201_CREATED)
def create_export_template(
    payload: ExportTemplateInput,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    name = payload.name.strip()
    if db.scalar(select(ExportTemplate).where(ExportTemplate.name == name)):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="同名导出模板已存在")
    fields, filters = _validate_export_template(payload)
    template = ExportTemplate(name=name, fields=fields, filters=filters, include_provenance=payload.include_provenance, mask_sensitive=payload.mask_sensitive, created_by_id=user.id)
    db.add(template)
    db.flush()
    _record_template_revision(db, "export", template.id, _serialize_export_template(template), "created", user)
    audit(db, "create_export_template", "export_template", template.id, actor=user, after=_serialize_export_template(template), request=request)
    db.commit()
    return _serialize_export_template(template)


@app.put("/api/export-templates/{template_id}")
def update_export_template(
    template_id: int,
    payload: ExportTemplateInput,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    template = db.get(ExportTemplate, template_id)
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导出模板不存在")
    name = payload.name.strip()
    owner = db.scalar(select(ExportTemplate).where(ExportTemplate.name == name, ExportTemplate.id != template.id))
    if owner:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="同名导出模板已存在")
    fields, filters = _validate_export_template(payload)
    before = _serialize_export_template(template)
    template.name, template.fields, template.filters = name, fields, filters
    template.include_provenance, template.mask_sensitive = payload.include_provenance, payload.mask_sensitive
    db.flush()
    _record_template_revision(db, "export", template.id, _serialize_export_template(template), "updated", user)
    audit(db, "update_export_template", "export_template", template.id, actor=user, before=before, after=_serialize_export_template(template), request=request)
    db.commit()
    return _serialize_export_template(template)


@app.delete("/api/export-templates/{template_id}")
def delete_export_template(
    template_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, bool]:
    require_csrf(request)
    template = db.get(ExportTemplate, template_id)
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导出模板不存在")
    snapshot = _serialize_export_template(template)
    _record_template_revision(db, "export", template.id, snapshot, "deleted", user)
    audit(db, "delete_export_template", "export_template", template.id, actor=user, before=snapshot, after={"deleted": True}, request=request)
    db.delete(template)
    db.commit()
    return {"ok": True}


@app.post("/api/export-templates/{template_id}/export")
def export_with_template(
    template_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    _require_capability(user, "student_export", "导出")
    template = db.get(ExportTemplate, template_id)
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导出模板不存在")
    actor_id = user.id
    scope = _student_scope(db, user)
    def worker(task_id: str, worker_db: Session) -> dict[str, Any]:
        update_task(worker_db, task_id, 20, "正在按导出模板读取档案")
        target = create_student_export(worker_db, filters=template.filters or {}, fields=template.fields or None, include_provenance=template.include_provenance, scope=scope, mask_sensitive=bool(template.mask_sensitive))
        update_task(worker_db, task_id, 90, "正在写入 XLSX 文件")
        actor = worker_db.get(User, actor_id)
        audit(worker_db, "export_xlsx_template", "export_template", template.id, actor=actor, after={"template": template.name, "filename": target.name, "filters": template.filters or {}, "mask_sensitive": bool(template.mask_sensitive), "background_task": task_id})
        worker_db.commit()
        return {"download_url": f"/api/exports/{target.name}", "filename": target.name}

    task = submit_task(db, "xlsx_template_export", user, worker)
    audit(db, "queue_export_template", "background_task", task.id, actor=user, after={"template_id": template.id, "template": template.name}, request=request)
    db.commit()
    return {"task": serialize_task(task)}


@app.post("/api/quality-scans")
def create_quality_scan(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    _require_capability(user, "quality_manage", "数据质量管理")
    require_csrf(request)
    scan = run_quality_scan(db, user)
    audit(db, "run_quality_scan", "quality_scan", scan.id, actor=user, after=scan.summary, request=request)
    db.commit()
    return serialize_quality_scan(scan)


@app.get("/api/quality-scans/latest")
def latest_quality_scan(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    _require_capability(user, "quality_manage", "数据质量管理")
    scan = db.scalar(select(QualityScan).order_by(QualityScan.created_at.desc()).limit(1))
    if scan is None:
        scan = run_quality_scan(db, None)
        db.commit()
    return serialize_quality_scan(scan)


@app.put("/api/quality-issues/{case_id}")
def update_quality_issue(
    case_id: int,
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    _require_capability(user, "quality_manage", "数据质量管理")
    require_csrf(request)
    case = db.get(QualityIssueCase, case_id)
    if not case:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="质量问题不存在")
    status_value = str(payload.get("status") or "").strip()
    if status_value not in {"open", "resolved", "ignored"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="问题状态不正确")
    assignee_id = payload.get("assignee_id")
    if assignee_id is not None and not db.get(User, int(assignee_id)):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="处理人不存在")
    before = {"status": case.status, "assignee_id": case.assignee_id, "resolution_note": case.resolution_note}
    case.status = status_value
    case.assignee_id = int(assignee_id) if assignee_id is not None else None
    case.resolution_note = str(payload.get("resolution_note") or "").strip()[:1000] or None
    case.resolved_at = utcnow() if status_value in {"resolved", "ignored"} else None
    audit(db, "update_quality_issue", "quality_issue_case", case.id, actor=user, before=before, after={"status": case.status, "assignee_id": case.assignee_id, "resolution_note": case.resolution_note}, request=request)
    db.commit()
    return {"id": case.id, "status": case.status, "assignee_id": case.assignee_id, "resolution_note": case.resolution_note, "resolved_at": case.resolved_at}


@app.get("/api/tasks")
def list_tasks(
    limit: int = 50,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[dict[str, Any]]:
    statement = select(BackgroundTask)
    if user.role not in {Role.SUPER_ADMIN, Role.ADMIN}:
        statement = statement.where(BackgroundTask.created_by_id == user.id)
    return [serialize_task(item) for item in db.scalars(statement.order_by(BackgroundTask.created_at.desc()).limit(min(max(limit, 1), 100)))]


@app.get("/api/tasks/{task_id}")
def get_task(task_id: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict[str, Any]:
    task = db.get(BackgroundTask, task_id)
    if not task or (user.role not in {Role.SUPER_ADMIN, Role.ADMIN} and task.created_by_id != user.id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="任务不存在")
    return serialize_task(task)


@app.get("/api/system/monitoring")
def system_monitoring(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    probe_started = time.perf_counter()
    snapshot = system_snapshot(db)
    gpu: dict[str, Any] = {"available": False}
    executable = shutil.which("nvidia-smi")
    if executable:
        try:
            result = subprocess.run([executable, "--query-gpu=name,memory.total,memory.used,temperature.gpu", "--format=csv,noheader,nounits"], capture_output=True, text=True, timeout=5, check=False)
            if result.returncode == 0:
                gpu = {"available": True, "gpus": [line.split(", ") for line in result.stdout.splitlines() if line.strip()]}
        except OSError:
            pass
    ai_status = snapshot["ai"]
    ai_status["health_probe_ms"] = round((time.perf_counter() - probe_started) * 1000, 1)
    return {"disk": snapshot["disk"], "database": snapshot["database"], "latest_backup": _serialize_backup(snapshot["latest_backup"]) if snapshot["latest_backup"] else None, "ai": ai_status, "gpu": gpu, "tasks": snapshot["tasks"], "alerts": [serialize_alert(item) for item in db.scalars(select(SystemAlert).where(SystemAlert.status.in_(["open", "acknowledged"])).order_by(SystemAlert.last_seen_at.desc()))]}


@app.get("/api/system/alerts")
def system_alerts(
    include_resolved: bool = False,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> list[dict[str, Any]]:
    statement = select(SystemAlert)
    if not include_resolved:
        statement = statement.where(SystemAlert.status.in_(["open", "acknowledged"]))
    return [serialize_alert(item) for item in db.scalars(statement.order_by(SystemAlert.last_seen_at.desc()).limit(200))]


@app.post("/api/system/alerts/check")
def check_system_alerts(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> list[dict[str, Any]]:
    require_csrf(request)
    alerts = evaluate_alerts(db, user)
    audit(db, "check_system_alerts", "system_alert", "monitoring", actor=user, after={"open": len(alerts)}, request=request)
    db.commit()
    return [serialize_alert(item) for item in alerts]


@app.post("/api/system/alerts/{alert_id}/acknowledge")
def acknowledge_system_alert(
    alert_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    alert = db.get(SystemAlert, alert_id)
    if not alert:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="告警不存在")
    if alert.status == "open":
        alert.status = "acknowledged"
        alert.acknowledged_at = utcnow()
        alert.acknowledged_by_id = user.id
    audit(db, "acknowledge_system_alert", "system_alert", alert.id, actor=user, request=request)
    db.commit()
    return serialize_alert(alert)


@app.get("/api/system/ai-evaluations/latest")
def latest_ai_evaluation(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any] | None:
    return serialize_ai_evaluation(db.scalar(select(AiEvaluationRun).order_by(AiEvaluationRun.created_at.desc()).limit(1)))


@app.post("/api/system/ai-evaluations")
def start_ai_evaluation(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    actor_id = user.id

    def worker(task_id: str, worker_db: Session) -> dict[str, Any]:
        update_task(worker_db, task_id, 15, "正在运行 AI 只读回归用例")
        actor = worker_db.get(User, actor_id)
        run = run_ai_regression(worker_db, actor)
        update_task(worker_db, task_id, 95, "正在保存 AI 回归结果")
        audit(worker_db, "run_ai_regression", "ai_evaluation", run.id, actor=actor, after=run.summary or {})
        worker_db.commit()
        return {"evaluation_id": run.id, "status": run.status}

    task = submit_task(db, "ai_regression", user, worker)
    audit(db, "queue_ai_regression", "background_task", task.id, actor=user, request=request)
    db.commit()
    return {"task": serialize_task(task)}


@app.get("/api/dashboard")
def dashboard(db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict[str, Any]:
    total = db.scalar(select(func.count()).select_from(Student)) or 0
    contactable = db.scalar(
        select(func.count()).select_from(Student).where(or_(Student.mobile_phone.is_not(None), Student.electronic_email.is_not(None)))
    ) or 0
    batch_statement = select(ImportBatch)
    if user.role == Role.TEACHER:
        batch_statement = batch_statement.where(ImportBatch.imported_by_id == user.id)
    batches = db.scalar(select(func.count()).select_from(batch_statement.subquery())) or 0
    pending_statement = (
        select(RelatedInfoCandidate)
        .join(ImportBatch, RelatedInfoCandidate.import_batch_id == ImportBatch.id)
        .where(RelatedInfoCandidate.status == CandidateStatus.PENDING)
    )
    if user.role == Role.TEACHER:
        pending_statement = pending_statement.where(ImportBatch.imported_by_id == user.id)
    pending = db.scalar(select(func.count()).select_from(pending_statement.subquery())) or 0
    latest = list(db.scalars(batch_statement.order_by(ImportBatch.created_at.desc()).limit(5)))
    return {
        "metrics": {"total_students": total, "contactable_students": contactable, "import_batches": batches, "pending_candidates": pending},
        "latest_imports": [
            {
                "id": batch.id,
                "status": batch.status.value,
                "total_rows": batch.total_rows,
                "created_rows": batch.created_rows,
                "updated_rows": batch.updated_rows,
                "created_at": batch.created_at,
            }
            for batch in latest
        ],
    }


@app.get("/api/students")
def students(
    keyword: str | None = None,
    current_class: str | None = None,
    school_major: str | None = None,
    college: str | None = None,
    school: str | None = None,
    gender: str | None = None,
    political_status: str | None = None,
    education_level: str | None = None,
    study_mode: str | None = None,
    page: int = 1,
    page_size: int = 50,
    sort_by: str = "student_no",
    sort_direction: str = "asc",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    if page < 1 or page_size < 10 or page_size > 100:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="分页参数不正确")
    if sort_direction not in {"asc", "desc"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="排序方向不正确")
    filters = {
        field: value
        for field, value in {
            "gender": gender,
            "political_status": political_status,
            "education_level": education_level,
            "study_mode": study_mode,
        }.items()
        if value and value.strip()
    }
    rows, total = list_students_page(
        db,
        keyword,
        current_class,
        school_major,
        college,
        school,
        filters,
        page,
        page_size,
        sort_by,
        sort_direction,
        _student_scope(db, user),
    )
    total_pages = max(1, (total + page_size - 1) // page_size)
    if page > total_pages:
        page = total_pages
        rows, total = list_students_page(
            db,
            keyword,
            current_class,
            school_major,
            college,
            school,
            filters,
            page,
            page_size,
            sort_by,
            sort_direction,
            _student_scope(db, user),
        )
    return {
        "items": [_serialize_student_for_user(db, user, student) for student in rows],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
    }


@app.get("/api/students/filter-options")
def student_filter_options(
    school: str | None = None,
    college: str | None = None,
    school_major: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, list[str]]:
    return list_student_filter_options(
        db,
        school=school.strip() if school else None,
        college=college.strip() if college else None,
        school_major=school_major.strip() if school_major else None,
        scope=_student_scope(db, user),
    )


@app.get("/api/student-filters")
def saved_student_filters(db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> list[dict[str, Any]]:
    rows = list(db.scalars(select(SavedStudentFilter).where(SavedStudentFilter.user_id == user.id).order_by(SavedStudentFilter.updated_at.desc(), SavedStudentFilter.id.desc()).limit(50)))
    return [{"id": item.id, "name": item.name, "filters": item.filters or {}, "updated_at": item.updated_at} for item in rows]


@app.post("/api/student-filters")
def save_student_filter(
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    name = str(payload.get("name") or "").strip()
    raw_filters = payload.get("filters") or {}
    if not name or len(name) > 128 or not isinstance(raw_filters, dict):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="筛选名称或筛选条件无效")
    allowed = {"keyword", "school", "college", "school_major", "current_class", "gender", "political_status", "sort_by", "sort_direction"}
    filters = {key: str(value).strip() for key, value in raw_filters.items() if key in allowed and value is not None and str(value).strip()}
    record = db.scalar(select(SavedStudentFilter).where(SavedStudentFilter.user_id == user.id, SavedStudentFilter.name == name))
    before = {"name": record.name, "filters": record.filters or {}} if record else None
    if record:
        record.filters = filters
    else:
        record = SavedStudentFilter(user_id=user.id, name=name, filters=filters)
        db.add(record)
    db.flush()
    audit(db, "save_student_filter", "saved_student_filter", record.id, actor=user, before=before, after={"name": record.name, "filters": filters}, request=request)
    db.commit()
    return {"id": record.id, "name": record.name, "filters": record.filters or {}, "updated_at": record.updated_at}


@app.delete("/api/student-filters/{filter_id}")
def delete_saved_student_filter(
    filter_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, bool]:
    require_csrf(request)
    record = db.get(SavedStudentFilter, filter_id)
    if not record or record.user_id != user.id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="常用筛选不存在")
    audit(db, "delete_student_filter", "saved_student_filter", record.id, actor=user, before={"name": record.name, "filters": record.filters or {}}, request=request)
    db.delete(record)
    db.commit()
    return {"ok": True}


@app.get("/api/students/duplicates")
def duplicate_students(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> list[dict[str, Any]]:
    groups = list_duplicate_groups(db)
    scope = _student_scope(db, user)
    if not scope:
        return groups
    visible: list[dict[str, Any]] = []
    for group in groups:
        students = []
        for item in group["students"]:
            student = db.get(Student, item["id"])
            if student:
                try:
                    ensure_student_scope(db, user, student)
                    students.append(item)
                except HTTPException:
                    continue
        if len(students) > 1:
            visible.append(group | {"count": len(students), "students": students})
    return visible


@app.post("/api/students/{source_student_id}/merge")
def merge_duplicate_students(
    source_student_id: int,
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    _require_capability(user, "student_edit", "学生档案编辑")
    require_csrf(request)
    if str(payload.get("confirmation_phrase") or "").strip() != "合并学生":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请输入“合并学生”确认")
    try:
        target_student_id = int(payload.get("target_student_id"))
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请选择保留的目标学生") from exc
    source = db.get(Student, source_student_id)
    target = db.get(Student, target_student_id)
    if not source or not target:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="要合并的学生不存在")
    ensure_student_scope(db, user, source)
    ensure_student_scope(db, user, target)
    result = merge_students(db, source, target, user)
    audit(db, "merge_students", "student_merge", result["merge_id"], actor=user, before={"source": result["source"], "target": result["target_before"]}, after={"target": result["target_after"], "merged_fields": result["merged_fields"]}, request=request)
    db.commit()
    return {"ok": True, "message": "学生档案已合并，原学生快照已保留在合并记录中", **result}


@app.post("/api/students", status_code=status.HTTP_201_CREATED)
def add_student(
    payload: StudentCreate,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    _require_edit_access(user)
    ensure_new_student_scope(db, user, payload.model_dump())
    student = create_student(db, payload, user)
    audit(db, "create", "student", student.id, actor=user, after=student_to_dict(student), request=request)
    db.commit()
    return _serialize_student_for_user(db, user, student)


@app.patch("/api/students/{student_id}")
def edit_student(
    student_id: int,
    payload: StudentUpdate,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    _require_edit_access(user)
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生不存在")
    ensure_student_scope(db, user, student)
    before, after = update_student(db, student, payload, user)
    for field in payload.model_dump(exclude={"row_version"}, exclude_unset=True):
        db.add(FieldProvenance(student_id=student.id, field_name=field, source_locator="平台编辑", raw_value=str(after.get(field) or ""), confidence=100))
    audit(db, "update", "student", student.id, actor=user, before=before, after=after, request=request)
    db.commit()
    return _serialize_student_for_user(db, user, student)


@app.post("/api/students/bulk-update")
def bulk_edit_students(
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """Apply the same low-risk profile fields to selected students.

    Each changed student receives its own version and audit event so the usual
    audit rollback protections continue to apply after a batch operation.
    """
    require_csrf(request)
    _require_edit_access(user)
    raw_ids = payload.get("student_ids")
    raw_changes = payload.get("changes")
    if not isinstance(raw_ids, list) or not isinstance(raw_changes, dict):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="批量编辑参数不正确")
    try:
        student_ids = list(dict.fromkeys(int(value) for value in raw_ids))
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="学生选择列表格式不正确") from exc
    if not student_ids or len(student_ids) > 500:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="每次请选择 1 至 500 名学生")
    allowed_fields = {"school", "college", "school_major", "current_class", "political_status", "education_level", "study_mode", "remarks"}
    changes = {
        field: (None if value is None else str(value).strip() or None)
        for field, value in raw_changes.items()
        if field in allowed_fields
    }
    if not changes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请至少选择一个可更新字段")
    students = list(db.scalars(select(Student).where(Student.id.in_(student_ids))))
    if len(students) != len(student_ids):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="选择中包含不存在的学生")
    for student in students:
        ensure_student_scope(db, user, student)
    updated = 0
    unchanged = 0
    for student in students:
        before = student_to_dict(student)
        changed_fields = [field for field, value in changes.items() if before.get(field) != value]
        if not changed_fields:
            unchanged += 1
            continue
        for field in changed_fields:
            setattr(student, field, changes[field])
        student.row_version += 1
        db.flush()
        after = student_to_dict(student)
        record_student_version(db, student, user, changed_fields)
        for field in changed_fields:
            db.add(FieldProvenance(student_id=student.id, field_name=field, source_locator="平台批量编辑", raw_value=str(after.get(field) or ""), confidence=100))
        audit(db, "update", "student", student.id, actor=user, before=before, after=after, request=request)
        updated += 1
    audit(db, "bulk_update_students", "student_batch", ",".join(str(item) for item in student_ids), actor=user, after={"student_count": len(student_ids), "updated": updated, "unchanged": unchanged, "fields": sorted(changes)}, request=request)
    db.commit()
    return {"updated": updated, "unchanged": unchanged, "fields": sorted(changes)}


@app.delete("/api/students/{student_id}")
def delete_student(
    student_id: int,
    payload: StudentDeletion,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, bool]:
    require_csrf(request)
    _require_edit_access(user)
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生不存在")
    ensure_student_scope(db, user, student)
    if payload.student_no != student.student_no or payload.confirmation_phrase != "永久删除":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="删除确认信息不正确")
    before = student_to_dict(student)
    recycle_record = permanently_delete_student(db, student, user)
    audit(
        db,
        "delete_student",
        "student",
        student_id,
        actor=user,
        before=before,
        after={"student_no": before["student_no"], "full_name": before["full_name"], "deleted": True, "recycle_id": recycle_record.id, "expires_at": recycle_record.expires_at},
        request=request,
    )
    db.commit()
    return {"ok": True}


@app.get("/api/students/{student_id}/versions")
def student_versions(student_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> list[dict[str, Any]]:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生不存在")
    ensure_student_scope(db, user, student)
    versions = list_student_versions(db, student_id)
    users = {item.id: item for item in db.scalars(select(User).where(User.id.in_([version.changed_by_id for version in versions if version.changed_by_id])))} if versions else {}
    return [{"id": version.id, "version_no": version.version_no, "changed_fields": version.changed_fields or [], "changed_by": (users[version.changed_by_id].display_name or users[version.changed_by_id].username) if version.changed_by_id in users else "系统", "created_at": version.created_at, "snapshot": serialize_student_for_user(db, user, version.snapshot or {})} for version in versions]


@app.get("/api/students/{student_id}/timeline")
def get_student_timeline(student_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> list[dict[str, Any]]:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生不存在")
    ensure_student_scope(db, user, student)
    return student_timeline(db, student)


@app.get("/api/student-reminders")
def student_reminders(
    limit: int = 100,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN, Role.TEACHER)),
) -> list[dict[str, Any]]:
    return list_student_reminders(db, _student_scope(db, user), limit)


@app.post("/api/students/{student_id}/versions/{version_id}/restore")
def restore_version(
    student_id: int,
    version_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    if user.role not in {Role.SUPER_ADMIN, Role.ADMIN}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="只有管理员可以恢复学生历史版本")
    student = db.get(Student, student_id)
    version = db.get(StudentVersion, version_id)
    if not student or not version or version.student_id != student_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生或历史版本不存在")
    ensure_student_scope(db, user, student)
    before, after = restore_student_version(db, student, version, user)
    audit(db, "restore_student_version", "student", student.id, actor=user, before=before, after={"restored_from_version": version.id, "student": after}, request=request)
    db.commit()
    return _serialize_student_for_user(db, user, student)


def _serialize_deleted_student(record: DeletedStudent) -> dict[str, Any]:
    return {"id": record.id, "student_no": record.student_no, "full_name": record.full_name, "deleted_at": record.deleted_at, "expires_at": record.expires_at, "restored_at": record.restored_at, "related_card_count": len(record.related_cards or [])}


@app.get("/api/recycle-bin")
def recycle_bin(
    include_restored: bool = False,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> list[dict[str, Any]]:
    return [_serialize_deleted_student(record) for record in list_deleted_students(db, include_restored)]


@app.post("/api/recycle-bin/{record_id}/restore")
def restore_recycled_student(
    record_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    record = db.get(DeletedStudent, record_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="回收站记录不存在")
    student = restore_deleted_student(db, record, user)
    audit(db, "restore_deleted_student", "deleted_student", record.id, actor=user, after={"student_id": student.id, "student_no": student.student_no}, request=request)
    db.commit()
    return _serialize_student_for_user(db, user, student)


@app.delete("/api/recycle-bin/{record_id}")
def purge_recycled_student(
    record_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, bool]:
    require_csrf(request)
    record = db.get(DeletedStudent, record_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="回收站记录不存在")
    audit(db, "purge_deleted_student", "deleted_student", record.id, actor=user, before=_serialize_deleted_student(record), after={"purged": True}, request=request)
    db.delete(record)
    db.commit()
    return {"ok": True}


@app.get("/api/students/{student_id}/provenance")
def student_provenance(student_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> list[dict[str, Any]]:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生不存在")
    ensure_student_scope(db, user, student)
    rows = get_provenance(db, student_id)
    document_ids = {row.source_document_id for row in rows if row.source_document_id}
    documents = {document.id: document for document in db.scalars(select(SourceDocument).where(SourceDocument.id.in_(document_ids)))} if document_ids else {}
    return [
        {
            "field_name": row.field_name,
            "file": documents[row.source_document_id].original_filename if row.source_document_id in documents else "平台编辑",
            "document_id": row.source_document_id,
            "document_available": bool(documents.get(row.source_document_id) and documents[row.source_document_id].status != "deleted"),
            "sheet": row.source_sheet,
            "row": row.source_row,
            "column": row.source_column,
            "locator": row.source_locator,
            "raw_value": serialize_student_for_user(db, user, {row.field_name: row.raw_value}).get(row.field_name),
            "confidence": row.confidence,
            "recorded_at": row.recorded_at,
        }
        for row in rows
    ]


@app.get("/api/students/{student_id}/related-info-cards")
def student_related_info_cards(student_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> list[dict[str, Any]]:
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生不存在")
    ensure_student_scope(db, user, student)
    cards = list(
        db.scalars(
            select(StudentRelatedInfoCard)
            .where(StudentRelatedInfoCard.student_id == student_id)
            .order_by(StudentRelatedInfoCard.imported_at.desc(), StudentRelatedInfoCard.id.desc())
        )
    )
    document_ids = {card.source_document_id for card in cards}
    importer_ids = {card.imported_by_id for card in cards}
    documents = {document.id: document for document in db.scalars(select(SourceDocument).where(SourceDocument.id.in_(document_ids)))} if document_ids else {}
    importers = {importer.id: importer for importer in db.scalars(select(User).where(User.id.in_(importer_ids)))} if importer_ids else {}
    return [
        {
            "id": card.id,
            "title": card.title,
            "payload": card.excel_payload,
            "source_file": documents[card.source_document_id].original_filename if card.source_document_id in documents else card.title,
            "source_document_id": card.source_document_id,
            "source_available": bool(documents.get(card.source_document_id) and documents[card.source_document_id].status != "deleted"),
            "imported_at": card.imported_at,
            "imported_by": (importers[card.imported_by_id].display_name or importers[card.imported_by_id].username) if card.imported_by_id in importers else "已删除用户",
        }
        for card in cards
    ]


@app.delete("/api/students/{student_id}/related-info-cards/{card_id}")
def delete_student_related_info_card(
    student_id: int,
    card_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, bool]:
    require_csrf(request)
    _require_edit_access(user)
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生不存在")
    ensure_student_scope(db, user, student)
    card = db.get(StudentRelatedInfoCard, card_id)
    if not card or card.student_id != student.id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生相关信息词条不存在")
    card_snapshot = {"student_id": card.student_id, "source_document_id": card.source_document_id, "import_batch_id": card.import_batch_id, "imported_by_id": card.imported_by_id, "title": card.title, "excel_payload": card.excel_payload or {}, "imported_at": card.imported_at}
    audit(
        db,
        "delete_related_info_card",
        "student_related_info_card",
        card.id,
        actor=user,
        before=card_snapshot | {"student_no": student.student_no},
        after={"deleted": True},
        request=request,
    )
    db.delete(card)
    db.commit()
    return {"ok": True}


@app.post("/api/imports/excel")
async def upload_excel(
    request: Request,
    file: UploadFile = File(...),
    mode: str = Form("upsert"),
    required_fields_json: str | None = Form(None),
    update_policy: str = Form("overwrite"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    _require_edit_access(user)
    try:
        required_fields = json.loads(required_fields_json) if required_fields_json else []
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="必填字段格式不正确") from exc
    if not isinstance(required_fields, list):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="必填字段格式不正确")
    document, content = await register_upload(db, file, user)
    if document.file_type != "excel":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请上传 Excel 文件")
    try:
        batch = import_excel(db, document, content, user, mode, required_fields=[str(field) for field in required_fields], update_policy=update_policy)
        audit(db, "import_excel", "import_batch", batch.id, actor=user, after={"document": document.original_filename, "mode": mode}, request=request)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {
        "batch_id": batch.id,
        "status": batch.status.value,
        "total_rows": batch.total_rows,
        "created_rows": batch.created_rows,
        "updated_rows": batch.updated_rows,
        "skipped_rows": batch.skipped_rows,
        "error_rows": batch.error_rows,
        "errors": batch.errors,
    }


@app.post("/api/imports/excel/preview")
async def preview_excel(
    request: Request,
    file: UploadFile = File(...),
    mode: str = Form("upsert"),
    mapping_json: str | None = Form(None),
    required_fields_json: str | None = Form(None),
    update_policy: str = Form("overwrite"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    _require_edit_access(user)
    try:
        requested_mapping = json.loads(mapping_json) if mapping_json else None
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="字段映射格式不正确") from exc
    if requested_mapping is not None and not isinstance(requested_mapping, dict):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="字段映射格式不正确")
    try:
        required_fields = json.loads(required_fields_json) if required_fields_json else []
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="必填字段格式不正确") from exc
    if not isinstance(required_fields, list):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="必填字段格式不正确")
    document, content = await register_upload(db, file, user)
    if document.file_type != "excel":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请上传 Excel 文件")
    try:
        preview_data, mapping = preview_excel_import(db, document, content, mode, requested_mapping, [str(field) for field in required_fields], update_policy)
        preview = ImportPreview(
            id=str(uuid4()),
            source_document_id=document.id,
            created_by_id=user.id,
            mode=mode,
            mapping={get_column_letter(column): field for column, field in mapping.items()},
            preview_data=preview_data,
        )
        db.add(preview)
        audit(
            db,
            "preview_excel_import",
            "import_preview",
            preview.id,
            actor=user,
            after={"document": document.original_filename, "mode": mode, "total_rows": preview_data["total_rows"], "conflict_rows": preview_data["conflict_rows"], "required_fields": preview_data["required_fields"], "update_policy": preview_data["update_policy"]},
            request=request,
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {**preview_data, "preview_id": preview.id, "mode": mode, "filename": document.original_filename}


@app.post("/api/imports/excel/commit")
def commit_excel_import(
    payload: ExcelImportCommit,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    _require_edit_access(user)
    preview = db.get(ImportPreview, str(payload.preview_id))
    if not preview or preview.created_by_id != user.id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入预检不存在或已失效")
    if preview.status != "pending":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该预检已经执行过导入")
    if preview.mode != payload.mode:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="导入方式已变化，请重新预检")
    if sorted(payload.required_fields) != sorted((preview.preview_data or {}).get("required_fields") or []) or payload.update_policy != (preview.preview_data or {}).get("update_policy", "overwrite"):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="导入策略已变化，请重新预检后再确认导入")
    normalized_mapping = {str(column).upper(): field for column, field in payload.mapping.items() if field}
    if normalized_mapping != (preview.mapping or {}):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="字段映射已变化，请重新预检后再确认导入")
    document = db.get(SourceDocument, preview.source_document_id)
    if not document or document.file_type != "excel" or not document_path(document).is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="预检原始文件不存在")
    if payload.background_task:
        preview.status = "queued"
        actor_id = user.id
        preview_id = preview.id

        def worker(task_id: str, worker_db: Session) -> dict[str, Any]:
            queued_preview = worker_db.get(ImportPreview, preview_id)
            queued_document = worker_db.get(SourceDocument, preview.source_document_id)
            actor = worker_db.get(User, actor_id)
            if not queued_preview or not queued_document or not actor or not document_path(queued_document).is_file():
                raise RuntimeError("导入预检或原始文件不存在")
            update_task(worker_db, task_id, 15, "正在读取 Excel")
            batch = import_excel(worker_db, queued_document, document_path(queued_document).read_bytes(), actor, payload.mode, normalized_mapping, payload.required_fields, payload.update_policy)
            queued_preview.status = "applied"
            queued_preview.applied_batch_id = batch.id
            queued_preview.applied_at = utcnow()
            audit(worker_db, "commit_excel_import", "import_batch", batch.id, actor=actor, after={"preview_id": queued_preview.id, "document": queued_document.original_filename, "mode": payload.mode, "created": batch.created_rows, "updated": batch.updated_rows, "errors": batch.error_rows, "background_task": task_id})
            worker_db.commit()
            return {"batch_id": batch.id, "status": batch.status.value, "total_rows": batch.total_rows, "created_rows": batch.created_rows, "updated_rows": batch.updated_rows, "skipped_rows": batch.skipped_rows, "error_rows": batch.error_rows, "errors": batch.errors}

        task = submit_task(db, "excel_student_import", user, worker)
        audit(db, "queue_excel_import", "background_task", task.id, actor=user, after={"preview_id": preview.id, "document": document.original_filename}, request=request)
        db.commit()
        return {"task": serialize_task(task)}
    try:
        batch = import_excel(db, document, document_path(document).read_bytes(), user, payload.mode, normalized_mapping, payload.required_fields, payload.update_policy)
        preview.status = "applied"
        preview.applied_batch_id = batch.id
        preview.applied_at = utcnow()
        audit(
            db,
            "commit_excel_import",
            "import_batch",
            batch.id,
            actor=user,
            after={"preview_id": preview.id, "document": document.original_filename, "mode": payload.mode, "created": batch.created_rows, "updated": batch.updated_rows, "errors": batch.error_rows},
            request=request,
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {
        "batch_id": batch.id,
        "status": batch.status.value,
        "total_rows": batch.total_rows,
        "created_rows": batch.created_rows,
        "updated_rows": batch.updated_rows,
        "skipped_rows": batch.skipped_rows,
        "error_rows": batch.error_rows,
        "errors": batch.errors,
    }


@app.post("/api/imports/word")
async def upload_word(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    _require_capability(user, "related_review", "学生相关信息审核")
    _require_edit_access(user)
    document, content = await register_upload(db, file, user)
    if document.file_type != "word":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请上传 Word 文件")
    try:
        candidates = import_word_for_review(db, document, content, user)
        audit(db, "import_word", "source_document", document.id, actor=user, after={"candidates": len(candidates)}, request=request)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"document_id": document.id, "candidates": len(candidates), "message": "已生成候选资料，请在审核区确认后写入学生档案"}


@app.post("/api/imports/related-info")
async def upload_related_info(
    request: Request,
    file: UploadFile = File(...),
    background: bool = Form(False),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    _require_capability(user, "related_review", "学生相关信息审核")
    _require_edit_access(user)
    document, content = await register_upload(db, file, user)
    if document.file_type not in {"word", "excel"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请上传 Word 或 Excel 文件")
    if background:
        actor_id = user.id
        document_id = document.id

        def worker(task_id: str, worker_db: Session) -> dict[str, Any]:
            queued_document = worker_db.get(SourceDocument, document_id)
            actor = worker_db.get(User, actor_id)
            if not queued_document or not actor or not document_path(queued_document).is_file():
                raise RuntimeError("原始文件不存在")
            update_task(worker_db, task_id, 15, "正在读取原始文件")
            batch = import_related_info(worker_db, queued_document, document_path(queued_document).read_bytes(), actor)
            update_task(worker_db, task_id, 88, "正在生成审核记录")
            audit(worker_db, "import_related_info", "import_batch", batch.id, actor=actor, after={"document": queued_document.original_filename, "pending": batch.created_rows, "skipped": batch.skipped_rows, "errors": batch.error_rows, "background_task": task_id})
            worker_db.commit()
            return {"batch_id": batch.id, "status": batch.status.value, "total_rows": batch.total_rows, "created_rows": batch.created_rows, "updated_rows": batch.updated_rows, "skipped_rows": batch.skipped_rows, "error_rows": batch.error_rows, "errors": batch.errors}

        task = submit_task(db, "related_info_import", user, worker)
        audit(db, "queue_related_info_import", "background_task", task.id, actor=user, after={"document": document.original_filename, "file_type": document.file_type}, request=request)
        db.commit()
        return {"task": serialize_task(task)}
    try:
        batch = import_related_info(db, document, content, user)
        audit(
            db,
            "import_related_info",
            "import_batch",
            batch.id,
            actor=user,
            after={"document": document.original_filename, "pending": batch.created_rows, "skipped": batch.skipped_rows, "errors": batch.error_rows},
            request=request,
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {
        "batch_id": batch.id,
        "status": batch.status.value,
        "total_rows": batch.total_rows,
        "created_rows": batch.created_rows,
        "updated_rows": batch.updated_rows,
        "skipped_rows": batch.skipped_rows,
        "error_rows": batch.error_rows,
        "errors": batch.errors,
    }


def _import_batch_is_undoable(batch: ImportBatch) -> bool:
    return batch.rollback_status == "available" or (batch.mode == "related_info" and batch.rollback_status is None)


@app.get("/api/imports")
def imports(db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> list[dict[str, Any]]:
    batch_statement = select(ImportBatch)
    if user.role == Role.TEACHER:
        batch_statement = batch_statement.where(ImportBatch.imported_by_id == user.id)
    batches = list(db.scalars(batch_statement.order_by(ImportBatch.created_at.desc()).limit(100)))
    documents = {document.id: document for document in db.scalars(select(SourceDocument).where(SourceDocument.id.in_([batch.source_document_id for batch in batches])))} if batches else {}
    user_batches = list(db.scalars(select(ImportBatch).where(ImportBatch.imported_by_id == user.id).order_by(ImportBatch.created_at.desc(), ImportBatch.id.desc()).limit(100)))
    latest_undoable_batch = next((batch for batch in user_batches if _import_batch_is_undoable(batch)), None)
    return [
        {
            "id": batch.id,
            "document_id": batch.source_document_id,
            "filename": documents.get(batch.source_document_id).original_filename if batch.source_document_id in documents else "已删除文件",
            "document_available": bool(documents.get(batch.source_document_id) and documents[batch.source_document_id].status != "deleted"),
            "mode": batch.mode,
            "status": batch.status.value,
            "total_rows": batch.total_rows,
            "created_rows": batch.created_rows,
            "updated_rows": batch.updated_rows,
            "skipped_rows": batch.skipped_rows,
            "error_rows": batch.error_rows,
            "rollback_status": batch.rollback_status or ("available" if batch.mode == "related_info" else None),
            "rollback_created": batch.created_rows if batch.mode == "related_info" else len((batch.rollback_data or {}).get("created") or []),
            "rollback_updated": len((batch.rollback_data or {}).get("updated") or []),
            "can_undo_latest": bool(latest_undoable_batch and batch.id == latest_undoable_batch.id and may_edit(user)),
            "created_at": batch.created_at,
        }
        for batch in batches
    ]


@app.post("/api/imports/{batch_id}/rollback")
def rollback_student_import(
    batch_id: int,
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    _require_edit_access(user)
    if str(payload.get("confirmation_phrase") or "").strip() != "撤销导入":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请输入“撤销导入”确认")
    batch = db.get(ImportBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入批次不存在")
    if batch.imported_by_id != user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="只能撤销本人导入的数据")
    user_batches = list(db.scalars(select(ImportBatch).where(ImportBatch.imported_by_id == user.id).order_by(ImportBatch.created_at.desc(), ImportBatch.id.desc()).limit(100)))
    latest_undoable_batch = next((item for item in user_batches if _import_batch_is_undoable(item)), None)
    if not latest_undoable_batch or latest_undoable_batch.id != batch.id:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="只允许撤销本人最近一次可撤销的数据导入")
    scope = _student_scope(db, user)
    if scope:
        if batch.mode == "related_info":
            student_ids = set(db.scalars(select(RelatedInfoCandidate.student_id).where(RelatedInfoCandidate.import_batch_id == batch.id)))
            student_ids.update(db.scalars(select(StudentRelatedInfoCard.student_id).where(StudentRelatedInfoCard.import_batch_id == batch.id)))
        else:
            student_ids = {int(item.get("student_id") or 0) for group in (batch.rollback_data or {}).values() if isinstance(group, list) for item in group if isinstance(item, dict)}
        for student in db.scalars(select(Student).where(Student.id.in_(student_ids))):
            ensure_student_scope(db, user, student)
    result = rollback_related_info_batch(db, batch, user) if batch.mode == "related_info" else rollback_import_batch(db, batch, user)
    source_document = db.get(SourceDocument, batch.source_document_id)
    audit(
        db,
        "rollback_own_latest_import",
        "import_batch",
        batch.id,
        actor=user,
        after={"filename": source_document.original_filename if source_document else "已删除文件", **result},
        request=request,
    )
    db.commit()
    message = "导入批次已撤销" if not result["blocked"] else f"已部分撤销；{len(result['blocked'])} 名学生在导入后被修改，未自动覆盖"
    return {"ok": True, "message": message, **result}


@app.get("/api/imports/{batch_id}/report")
def import_report_preview(
    batch_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    batch = db.get(ImportBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入批次不存在")
    _require_teacher_import_access(user, batch.imported_by_id)
    document = db.get(SourceDocument, batch.source_document_id)
    importer = db.get(User, batch.imported_by_id)
    rollback_data = batch.rollback_data or {}
    rollback_status = batch.rollback_status or ("available" if batch.mode == "related_info" else "不支持")
    created_changes = [
        {"student_no": str(item.get("student_no") or "-"), "fields": "全部档案"}
        for item in (rollback_data.get("created") or [])[:30]
        if isinstance(item, dict)
    ]
    updated_changes = [
        {"student_no": str(item.get("student_no") or "-"), "fields": "、".join(str(field) for field in (item.get("changed_fields") or []) if field) or "-"}
        for item in (rollback_data.get("updated") or [])[:30]
        if isinstance(item, dict)
    ]
    errors = [
        {"row": item.get("row", "-"), "message": str(item.get("message") or "未说明原因")}
        for item in (batch.errors or [])[:80]
        if isinstance(item, dict)
    ]
    related_changes = [item for item in (rollback_data.get("related_changes") or []) if isinstance(item, dict)]
    if batch.mode == "related_info":
        created_changes = [{"student_no": str(item.get("student_no") or "-"), "fields": "Excel 信息词条" if item.get("kind") == "excel_card" else "学生备注"} for item in related_changes[:30]]
    return {
        "id": batch.id,
        "filename": document.original_filename if document else "已删除文件",
        "mode": batch.mode,
        "status": batch.status.value,
        "imported_by": (importer.display_name or importer.username) if importer else "已删除账号",
        "created_at": batch.created_at,
        "completed_at": batch.completed_at,
        "total_rows": batch.total_rows,
        "created_rows": batch.created_rows,
        "updated_rows": batch.updated_rows,
        "skipped_rows": batch.skipped_rows,
        "error_rows": batch.error_rows,
        "errors": errors,
        "errors_truncated": batch.error_rows > len(errors),
        "rollback_status": rollback_status,
        "rollback_created": batch.created_rows if batch.mode == "related_info" else len(rollback_data.get("created") or []),
        "rollback_updated": len(rollback_data.get("updated") or []),
        "rollback_changes": {"created": created_changes, "updated": updated_changes},
    }


@app.get("/api/imports/{batch_id}/errors.csv")
def download_import_errors(
    batch_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    batch = db.get(ImportBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入批次不存在")
    if batch.imported_by_id != user.id and user.role not in {Role.SUPER_ADMIN, Role.ADMIN}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="只能下载本人导入的错误明细")
    document = db.get(SourceDocument, batch.source_document_id)
    stream = io.StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow(["原始位置", "错误原因"])
    for item in batch.errors or []:
        if isinstance(item, dict):
            writer.writerow([str(item.get("row") or "-"), str(item.get("message") or "未说明原因")])
    filename_stem = Path(document.original_filename if document else f"import_{batch.id}").stem
    headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{filename_stem}_错误明细.csv"}
    return Response(content=("\ufeff" + stream.getvalue()).encode("utf-8"), media_type="text/csv; charset=utf-8", headers=headers)


@app.get("/api/imports/{batch_id}/errors.xlsx")
def download_import_error_rows(
    batch_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Create an editable workbook containing only failed source rows."""
    batch = db.get(ImportBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入批次不存在")
    if batch.imported_by_id != user.id and user.role not in {Role.SUPER_ADMIN, Role.ADMIN}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="只能下载本人导入的错误行")
    document = db.get(SourceDocument, batch.source_document_id)
    if not document or document.file_type != "excel" or not document_path(document).is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="原始 Excel 文件不存在")
    error_rows = {}
    for item in batch.errors or []:
        if isinstance(item, dict):
            try:
                error_rows[int(item.get("row"))] = str(item.get("message") or "未说明原因")
            except (TypeError, ValueError):
                continue
    if not error_rows:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="该批次没有可下载的错误行")
    try:
        workbook = load_workbook(io.BytesIO(document_path(document).read_bytes()), data_only=False)
        visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        worksheet = visible_sheets[0] if visible_sheets else workbook.worksheets[0]
        output = Workbook()
        target = output.active
        target.title = worksheet.title[:31] or "错误行"
        headers = [cell.value for cell in worksheet[1]]
        target.append([*headers, "导入错误"])
        for row_number in sorted(error_rows):
            if row_number > worksheet.max_row:
                continue
            target.append([cell.value for cell in worksheet[row_number]] + [error_rows[row_number]])
        stream = io.BytesIO()
        output.save(stream)
        workbook.close()
        filename_stem = Path(document.original_filename).stem
        headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{filename_stem}_错误行修正模板.xlsx"}
        return Response(content=stream.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"错误行模板生成失败: {exc}") from exc


@app.post("/api/imports/{batch_id}/retry-errors")
async def retry_import_errors(
    batch_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """Retry corrected rows as a new, fully auditable import batch."""
    require_csrf(request)
    _require_edit_access(user)
    batch = db.get(ImportBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入批次不存在")
    if batch.mode == "related_info":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="学生相关信息请先人工匹配后导入，不能用主档案错误行重试")
    if batch.imported_by_id != user.id and user.role not in {Role.SUPER_ADMIN, Role.ADMIN}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="只能重试本人导入的错误行")
    original = db.get(SourceDocument, batch.source_document_id)
    if not original:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="原始导入文件不存在")
    try:
        document, content = await register_upload(db, file, user)
    except Exception:
        db.rollback()
        raise
    if document.file_type != "excel":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="错误行修正模板必须是 Excel 文件")
    mapping = (batch.mapping or {}).get("columns") or batch.mapping or {}
    mapping = {str(column).upper(): field for column, field in mapping.items() if field}
    required_fields = [str(field) for field in ((batch.mapping or {}).get("required_fields") or [])]
    update_policy = str((batch.mapping or {}).get("update_policy") or "overwrite")
    try:
        retry_batch = import_excel(db, document, content, user, batch.mode, mapping, required_fields, update_policy)
        retry_batch.mapping = dict(retry_batch.mapping or {}) | {"retry_of_batch_id": batch.id}
        audit(
            db,
            "retry_import_errors",
            "import_batch",
            retry_batch.id,
            actor=user,
            after={"retry_of_batch_id": batch.id, "document": document.original_filename, "created": retry_batch.created_rows, "updated": retry_batch.updated_rows, "errors": retry_batch.error_rows},
            request=request,
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"batch_id": retry_batch.id, "status": retry_batch.status.value, "total_rows": retry_batch.total_rows, "created_rows": retry_batch.created_rows, "updated_rows": retry_batch.updated_rows, "skipped_rows": retry_batch.skipped_rows, "error_rows": retry_batch.error_rows, "errors": retry_batch.errors}


def _source_document_student_ids(db: Session, document_id: int) -> set[int]:
    student_ids = set(db.scalars(select(FieldProvenance.student_id).where(FieldProvenance.source_document_id == document_id)))
    student_ids.update(db.scalars(select(StudentRelatedInfoCard.student_id).where(StudentRelatedInfoCard.source_document_id == document_id)))
    student_ids.update(db.scalars(select(RelatedInfoCandidate.student_id).where(RelatedInfoCandidate.source_document_id == document_id)))
    for batch in db.scalars(select(ImportBatch).where(ImportBatch.source_document_id == document_id)):
        for group_name in ("created", "updated"):
            for item in (batch.rollback_data or {}).get(group_name) or []:
                if isinstance(item, dict) and item.get("student_id"):
                    student_ids.add(int(item["student_id"]))
    return {int(student_id) for student_id in student_ids if student_id}


def _source_document_student_rows(db: Session, document_id: int) -> dict[str, dict[int, int]]:
    rows: dict[str, dict[int, int]] = {}

    def remember(student_id: int | None, sheet: str | None, row: int | None) -> None:
        if student_id and sheet and row and row > 0:
            rows.setdefault(sheet, {})[row] = student_id

    for item in db.scalars(select(FieldProvenance).where(FieldProvenance.source_document_id == document_id)):
        remember(item.student_id, item.source_sheet, item.source_row)
    for item in db.scalars(select(StudentRelatedInfoCard).where(StudentRelatedInfoCard.source_document_id == document_id)):
        payload = item.excel_payload or {}
        source_row = payload.get("source_row")
        remember(item.student_id, str(payload.get("sheet_name") or "") or None, int(source_row) if source_row else None)
    for item in db.scalars(select(RelatedInfoCandidate).where(RelatedInfoCandidate.source_document_id == document_id)):
        remember(item.student_id, item.source_sheet, item.source_row)
    return rows


def _serialize_source_document(db: Session, document: SourceDocument) -> dict[str, Any]:
    import_count = db.scalar(select(func.count()).select_from(ImportBatch).where(ImportBatch.source_document_id == document.id)) or 0
    card_count = db.scalar(select(func.count()).select_from(StudentRelatedInfoCard).where(StudentRelatedInfoCard.source_document_id == document.id)) or 0
    return {"id": document.id, "filename": document.original_filename, "file_type": document.file_type, "size_bytes": document.size_bytes, "sha256": document.sha256, "version_group": document.version_group, "version_no": document.version_no, "tags": document.tags or [], "status": document.status, "uploaded_at": document.uploaded_at, "archived_at": document.archived_at, "import_count": import_count, "related_card_count": card_count, "associated_student_count": len(_source_document_student_ids(db, document.id))}


@app.get("/api/documents")
def source_documents(
    status_value: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> list[dict[str, Any]]:
    _require_capability(user, "source_manage", "原始资料管理")
    statement = select(SourceDocument)
    if status_value:
        statement = statement.where(SourceDocument.status == status_value)
    else:
        # Deleted source binaries stay as lineage records for imports and audits,
        # but are deliberately hidden from the material library.
        statement = statement.where(SourceDocument.status != "deleted")
    documents = list(db.scalars(statement.order_by(SourceDocument.uploaded_at.desc()).limit(500)))
    return [_serialize_source_document(db, document) for document in documents]


@app.put("/api/documents/{document_id}")
def update_source_document(
    document_id: int,
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    _require_capability(user, "source_manage", "原始资料管理")
    require_csrf(request)
    document = db.get(SourceDocument, document_id)
    if not document:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="原始文件不存在")
    if document.status == "deleted":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="原始文件已经删除，无法再编辑或恢复")
    before = {"tags": list(document.tags or []), "status": document.status}
    tags = payload.get("tags")
    if tags is not None:
        if not isinstance(tags, list):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="文件标签格式不正确")
        document.tags = list(dict.fromkeys(str(item).strip()[:32] for item in tags if str(item).strip()))[:20]
    requested_status = payload.get("status")
    if requested_status is not None:
        if requested_status not in {"active", "archived"}:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="文件状态不正确")
        document.status = requested_status
        document.archived_at = utcnow() if requested_status == "archived" else None
    after = {"tags": document.tags, "status": document.status}
    audit(db, "update_source_document", "source_document", document.id, actor=user, before=before, after=after, request=request)
    db.commit()
    return _serialize_source_document(db, document)


@app.delete("/api/documents/{document_id}")
def delete_source_document(
    document_id: int,
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    """Delete the original binary and remove its associated active student records."""
    _require_capability(user, "source_manage", "原始资料管理")
    require_csrf(request)
    if int(payload.get("confirmation_count") or 0) != 3 or str(payload.get("confirmation_phrase") or "").strip() != "永久删除原始资料":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="需要完成三次确认，并输入“永久删除原始资料”")
    document = db.get(SourceDocument, document_id)
    if not document:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="原始文件不存在")
    if document.status == "deleted":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="原始文件已经删除")

    before = _serialize_source_document(db, document)
    associated_student_ids = _source_document_student_ids(db, document.id)
    associated_students = list(db.scalars(select(Student).where(Student.id.in_(associated_student_ids)))) if associated_student_ids else []
    try:
        source_path = document_path(document)
        binary_deleted = source_path.is_file()
        if binary_deleted:
            source_path.unlink()
    except OSError as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="无法删除原始文件，请确认文件没有被其他程序占用") from exc

    document.status = "deleted"
    document.archived_at = utcnow()
    for student in associated_students:
        permanently_delete_student(db, student, user)

    batches = list(db.scalars(select(ImportBatch).where(ImportBatch.source_document_id == document.id)))
    for batch in batches:
        batch.rollback_status = "source_deleted"
        batch.rollback_data = {"source_deleted": True, "deleted_student_count": len(associated_students), "deleted_at": utcnow().isoformat()}

    source_match_reviews = list(db.scalars(select(ImportMatchReview).where(ImportMatchReview.source_document_id == document.id)))
    pending_matches = [review for review in source_match_reviews if review.status == "pending"]
    for review in source_match_reviews:
        review.status = "source_deleted"
        review.reviewed_by_id = user.id
        review.reviewed_at = utcnow()
        review.matched_student_id = None
        review.identity = {}
        review.payload = {}

    source_word_candidates = list(db.scalars(select(WordImportCandidate).where(WordImportCandidate.source_document_id == document.id)))
    for candidate in source_word_candidates:
        candidate.status = CandidateStatus.REJECTED
        candidate.reviewed_by_id = user.id
        candidate.reviewed_at = utcnow()
        candidate.candidate_data = {}
        candidate.evidence = []

    pending_previews = list(db.scalars(select(ImportPreview).where(ImportPreview.source_document_id == document.id, ImportPreview.status.in_(["pending", "queued"]))))
    for preview in pending_previews:
        preview.status = "source_deleted"
        preview.preview_data = {}

    audit(
        db,
        "delete_source_document",
        "source_document",
        document.id,
        actor=user,
        before=before,
        after={
            "deleted": True,
            "binary_deleted": binary_deleted,
            "deleted_student_count": len(associated_students),
            "deleted_student_nos": [student.student_no for student in associated_students[:100]],
            "ignored_match_review_count": len(pending_matches),
            "rejected_word_candidate_count": len(source_word_candidates),
            "cancelled_preview_count": len(pending_previews),
            "retained_lineage": {"import_count": before["import_count"], "related_card_count": before["related_card_count"]},
        },
        request=request,
    )
    db.commit()
    return {"ok": True, "filename": document.original_filename, "binary_deleted": binary_deleted, "deleted_students": len(associated_students)}


@app.delete("/api/documents/{document_id}/related-info-cards")
def delete_source_related_cards(
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, int]:
    _require_capability(user, "source_manage", "原始资料管理")
    require_csrf(request)
    document = db.get(SourceDocument, document_id)
    if not document:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="原始文件不存在")
    cards = list(db.scalars(select(StudentRelatedInfoCard).where(StudentRelatedInfoCard.source_document_id == document.id)))
    for card in cards:
        student = db.get(Student, card.student_id)
        if student:
            ensure_student_scope(db, user, student)
    card_snapshots = [{"id": card.id, "student_id": card.student_id, "source_document_id": card.source_document_id, "import_batch_id": card.import_batch_id, "imported_by_id": card.imported_by_id, "title": card.title, "excel_payload": card.excel_payload or {}, "imported_at": card.imported_at.isoformat() if card.imported_at else None} for card in cards]
    for card in cards:
        db.delete(card)
    audit(db, "delete_source_related_cards", "source_document", document.id, actor=user, before={"cards": card_snapshots}, after={"deleted_cards": len(cards)}, request=request)
    db.commit()
    return {"deleted_cards": len(cards)}


@app.get("/api/documents/{document_id}/download")
def download_document(
    document_id: int,
    request: Request,
    student_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    document = db.get(SourceDocument, document_id)
    if not document or document.status == "deleted" or not document_path(document).is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="原始文件不存在")
    is_source_manager = user.role in {Role.SUPER_ADMIN, Role.ADMIN} and has_capability(user, "source_manage")
    if document.uploaded_by_id == user.id or is_source_manager:
        audit(db, "download_source_document", "source_document", document.id, actor=user, after={"filename": document.original_filename, "student_scoped": False}, request=request)
        db.commit()
        return FileResponse(document_path(document), filename=document.original_filename, media_type=document.mime_type or "application/octet-stream")
    if not student_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="请从有权限的学生档案中下载该原始文件")
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生不存在")
    ensure_student_scope(db, user, student)
    associated_student_ids = _source_document_student_ids(db, document.id)
    if student.id not in associated_student_ids:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="该原始文件不包含当前学生的信息")
    related_students = list(db.scalars(select(Student).where(Student.id.in_(associated_student_ids))))
    try:
        content = create_student_scoped_source_copy(
            document_path(document).read_bytes(),
            document.file_type,
            student,
            related_students,
            _source_document_student_rows(db, document.id),
        )
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE, detail=str(exc)) from exc
    filename = sanitized_download_filename(document.original_filename, student.student_no)
    audit(
        db,
        "download_student_scoped_source_document",
        "source_document",
        document.id,
        actor=user,
        after={"filename": filename, "student_id": student.id, "student_no": student.student_no, "student_scoped": True},
        request=request,
    )
    db.commit()
    return StreamingResponse(
        io.BytesIO(content),
        media_type=document.mime_type or "application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@app.get("/api/word-candidates")
def word_candidates(db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> list[dict[str, Any]]:
    statement = select(WordImportCandidate).where(WordImportCandidate.status == CandidateStatus.PENDING)
    if user.role == Role.TEACHER:
        statement = statement.where(WordImportCandidate.created_by_id == user.id)
    records = list(db.scalars(statement.order_by(WordImportCandidate.created_at.desc()).limit(100)))
    documents = {document.id: document for document in db.scalars(select(SourceDocument).where(SourceDocument.id.in_([record.source_document_id for record in records])))} if records else {}
    return [
        {
            "id": record.id,
            "document_id": record.source_document_id,
            "filename": documents.get(record.source_document_id).original_filename if record.source_document_id in documents else "已删除文件",
            "data": record.candidate_data,
            "evidence": record.evidence,
            "confidence": record.confidence,
            "created_at": record.created_at,
        }
        for record in records
    ]


@app.get("/api/related-info-candidates")
def related_info_candidates(db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> list[dict[str, Any]]:
    _require_capability(user, "related_review", "学生相关信息审核")
    statement = (
        select(RelatedInfoCandidate)
        .join(ImportBatch, RelatedInfoCandidate.import_batch_id == ImportBatch.id)
        .where(RelatedInfoCandidate.status == CandidateStatus.PENDING)
    )
    if user.role == Role.TEACHER:
        statement = statement.where(ImportBatch.imported_by_id == user.id)
    records = list(db.scalars(statement.order_by(RelatedInfoCandidate.created_at.desc()).limit(200)))
    student_ids = {record.student_id for record in records}
    student_ids = _scope_student_ids(db, user, student_ids)
    records = [record for record in records if record.student_id in student_ids]
    document_ids = {record.source_document_id for record in records}
    students = {student.id: student for student in db.scalars(select(Student).where(Student.id.in_(student_ids)))} if student_ids else {}
    documents = {document.id: document for document in db.scalars(select(SourceDocument).where(SourceDocument.id.in_(document_ids)))} if document_ids else {}
    return [
        {
            "id": record.id,
            "student_id": record.student_id,
            "student_no": students[record.student_id].student_no if record.student_id in students else "已删除学生",
            "full_name": students[record.student_id].full_name if record.student_id in students else "已删除学生",
            "remarks": record.remarks,
            "content_type": record.content_type,
            "excel_payload": record.excel_payload if record.content_type == "excel_card" else None,
            "filename": documents[record.source_document_id].original_filename if record.source_document_id in documents else "已删除文件",
            "document_id": record.source_document_id,
            "sheet": record.source_sheet,
            "row": record.source_row,
            "locator": record.source_locator,
            "confidence": record.confidence,
            "created_at": record.created_at,
        }
        for record in records
    ]


def _validate_related_info_candidate_review(
    db: Session,
    user: User,
    candidate: RelatedInfoCandidate,
) -> tuple[ImportBatch, Student]:
    if candidate.status != CandidateStatus.PENDING:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该审核项已处理")
    batch = db.get(ImportBatch, candidate.import_batch_id)
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入批次不存在")
    _require_teacher_import_access(user, batch.imported_by_id)
    student = db.get(Student, candidate.student_id)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="关联学生不存在")
    ensure_student_scope(db, user, student)
    return batch, student


@app.post("/api/related-info-candidates/bulk-approve")
def bulk_approve_related_info_candidates(
    request: Request,
    payload: dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _require_capability(user, "related_review", "学生相关信息审核")
    require_csrf(request)
    raw_ids = payload.get("candidate_ids")
    if not isinstance(raw_ids, list):
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="请提交待审核项列表")
    candidate_ids: list[int] = []
    for raw_id in raw_ids:
        try:
            candidate_id = int(raw_id)
        except (TypeError, ValueError):
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="审核项编号格式不正确") from None
        if candidate_id > 0 and candidate_id not in candidate_ids:
            candidate_ids.append(candidate_id)
    if not candidate_ids:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="请至少选择一条待审核信息")
    if len(candidate_ids) > 200:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="一次最多确认 200 条审核信息")

    candidates_by_id = {
        candidate.id: candidate
        for candidate in db.scalars(select(RelatedInfoCandidate).where(RelatedInfoCandidate.id.in_(candidate_ids)))
    }
    missing_ids = [candidate_id for candidate_id in candidate_ids if candidate_id not in candidates_by_id]
    if missing_ids:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"审核项不存在：{missing_ids[0]}")
    candidates = [candidates_by_id[candidate_id] for candidate_id in candidate_ids]

    # Validate every item before writing anything, so a stale or unauthorized item cannot leave a partial batch behind.
    for candidate in candidates:
        _validate_related_info_candidate_review(db, user, candidate)

    try:
        approved = []
        for candidate in candidates:
            student = apply_related_info_candidate(db, candidate, user)
            audit(
                db,
                "approve_related_info",
                "related_info_candidate",
                candidate.id,
                actor=user,
                after={"student_id": student.id, "source_document_id": candidate.source_document_id, "bulk": True},
                request=request,
            )
            approved.append({"candidate_id": candidate.id, "student_id": student.id, "student_no": student.student_no})
        audit(
            db,
            "bulk_approve_related_info",
            "related_info_candidate_batch",
            ",".join(str(candidate_id) for candidate_id in candidate_ids),
            actor=user,
            after={"approved_count": len(approved), "items": approved},
            request=request,
        )
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="批量确认失败，未写入任何学生相关信息") from None
    return {"ok": True, "approved_count": len(approved), "candidate_ids": candidate_ids, "message": f"已确认 {len(approved)} 条学生相关信息"}


@app.post("/api/related-info-candidates/{candidate_id}/approve")
def approve_related_info_candidate(
    candidate_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _require_capability(user, "related_review", "学生相关信息审核")
    require_csrf(request)
    candidate = db.get(RelatedInfoCandidate, candidate_id)
    if not candidate:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="审核项不存在")
    _validate_related_info_candidate_review(db, user, candidate)
    student = apply_related_info_candidate(db, candidate, user)
    audit(db, "approve_related_info", "related_info_candidate", candidate.id, actor=user, after={"student_id": student.id, "source_document_id": candidate.source_document_id}, request=request)
    db.commit()
    return _serialize_student(student)


@app.post("/api/related-info-candidates/{candidate_id}/reject")
def reject_related_info_candidate(
    candidate_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, bool]:
    _require_capability(user, "related_review", "学生相关信息审核")
    require_csrf(request)
    candidate = db.get(RelatedInfoCandidate, candidate_id)
    if not candidate:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="审核项不存在")
    batch = db.get(ImportBatch, candidate.import_batch_id)
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入批次不存在")
    _require_teacher_import_access(user, batch.imported_by_id)
    if candidate.status != CandidateStatus.PENDING:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该审核项已处理")
    candidate.status = CandidateStatus.REJECTED
    candidate.reviewed_by_id = user.id
    candidate.reviewed_at = utcnow()
    audit(db, "reject_related_info", "related_info_candidate", candidate.id, actor=user, request=request)
    db.commit()
    return {"ok": True}


@app.get("/api/import-match-reviews")
def import_match_reviews(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[dict[str, Any]]:
    _require_capability(user, "related_review", "导入匹配审核")
    statement = select(ImportMatchReview).join(ImportBatch, ImportMatchReview.import_batch_id == ImportBatch.id).where(ImportMatchReview.status == "pending")
    if user.role == Role.TEACHER:
        statement = statement.where(ImportBatch.imported_by_id == user.id)
    records = list(
        db.scalars(
            statement.order_by(ImportMatchReview.created_at.desc())
            .limit(200)
        )
    )
    candidate_ids = {student_id for record in records for student_id in (record.candidate_student_ids or [])}
    documents = {document.id: document for document in db.scalars(select(SourceDocument).where(SourceDocument.id.in_({record.source_document_id for record in records})))} if records else {}
    students = {student.id: student for student in db.scalars(select(Student).where(Student.id.in_(candidate_ids)))} if candidate_ids else {}
    return [
        {
            "id": record.id,
            "document_id": record.source_document_id,
            "filename": documents.get(record.source_document_id).original_filename if record.source_document_id in documents else "已删除文件",
            "identity": record.identity or {},
            "payload": record.payload or {},
            "reason": record.match_reason,
            "created_at": record.created_at,
            "suggestions": [
                {"id": student.id, "student_no": student.student_no, "full_name": student.full_name, "school_major": student.school_major, "current_class": student.current_class}
                for student_id in (record.candidate_student_ids or [])
                if (student := students.get(student_id)) is not None
            ],
        }
        for record in records
    ]


@app.post("/api/import-match-reviews/{review_id}/match")
def match_import_review(
    review_id: int,
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _require_capability(user, "related_review", "导入匹配审核")
    require_csrf(request)
    review = db.get(ImportMatchReview, review_id)
    if not review:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="人工匹配项不存在")
    batch = db.get(ImportBatch, review.import_batch_id)
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入批次不存在")
    _require_teacher_import_access(user, batch.imported_by_id)
    try:
        student_id = int(payload.get("student_id"))
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请选择要关联的学生") from exc
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生不存在")
    ensure_student_scope(db, user, student)
    candidate = resolve_import_match_review(db, review, student, user)
    audit(db, "resolve_import_match", "import_match_review", review.id, actor=user, after={"student_id": student.id, "candidate_id": candidate.id}, request=request)
    db.commit()
    return {"id": candidate.id, "student_id": student.id, "student_no": student.student_no, "full_name": student.full_name}


@app.post("/api/import-match-reviews/{review_id}/ignore")
def ignore_import_review(
    review_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, bool]:
    _require_capability(user, "related_review", "导入匹配审核")
    require_csrf(request)
    review = db.get(ImportMatchReview, review_id)
    if not review:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="人工匹配项不存在")
    batch = db.get(ImportBatch, review.import_batch_id)
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入批次不存在")
    _require_teacher_import_access(user, batch.imported_by_id)
    dismiss_import_match_review(review, user)
    audit(db, "ignore_import_match", "import_match_review", review.id, actor=user, request=request)
    db.commit()
    return {"ok": True}


@app.post("/api/word-candidates/{candidate_id}/approve")
def approve_candidate(
    candidate_id: int,
    payload: CandidateApproval,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    _require_capability(user, "related_review", "学生相关信息审核")
    require_csrf(request)
    candidate = db.get(WordImportCandidate, candidate_id)
    if not candidate:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="候选项不存在")
    _require_teacher_import_access(user, candidate.created_by_id)
    student = apply_word_candidate(db, candidate, payload, user)
    candidate.reviewed_by_id = user.id
    candidate.reviewed_at = utcnow()
    audit(db, "approve_word_candidate", "word_import_candidate", candidate.id, actor=user, after={"student_id": student.id}, request=request)
    db.commit()
    return _serialize_student(student)


@app.post("/api/word-candidates/{candidate_id}/reject")
def reject_candidate(candidate_id: int, request: Request, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict[str, bool]:
    _require_capability(user, "related_review", "学生相关信息审核")
    require_csrf(request)
    candidate = db.get(WordImportCandidate, candidate_id)
    if not candidate:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="候选项不存在")
    _require_teacher_import_access(user, candidate.created_by_id)
    if candidate.status != CandidateStatus.PENDING:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该候选项已处理")
    candidate.status = CandidateStatus.REJECTED
    candidate.reviewed_by_id = user.id
    candidate.reviewed_at = utcnow()
    audit(db, "reject_word_candidate", "word_import_candidate", candidate.id, actor=user, request=request)
    db.commit()
    return {"ok": True}


@app.post("/api/exports")
def export_students(
    request: Request,
    payload: dict[str, Any] | None = Body(default=None),
    keyword: str | None = None,
    current_class: str | None = None,
    school_major: str | None = None,
    college: str | None = None,
    school: str | None = None,
    include_provenance: bool = True,
    mask_sensitive: bool = False,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    _require_capability(user, "student_export", "导出")
    filters = {key: value for key, value in {"keyword": keyword, "current_class": current_class, "school_major": school_major, "college": college, "school": school}.items() if value}
    actor_id = user.id
    scope = _student_scope(db, user)
    effective_mask_sensitive = bool(mask_sensitive)
    raw_student_ids = payload.get("student_ids") if payload and "student_ids" in payload else request.query_params.get("student_ids")
    selected_student_ids: list[int] | None = None
    if raw_student_ids is not None:
        values = raw_student_ids.split(",") if isinstance(raw_student_ids, str) else raw_student_ids
        if not isinstance(values, list):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="学生选择列表格式不正确")
        try:
            selected_student_ids = list(dict.fromkeys(int(value) for value in values if str(value).strip()))
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="学生选择列表格式不正确") from exc
        if len(selected_student_ids) > 10000:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="单次最多选择 10000 名学生导出")
        if not selected_student_ids:
            selected_student_ids = None
        else:
            accessible_ids = set(
                db.scalars(
                    build_student_query(scope=scope)
                    .with_only_columns(Student.id, maintain_column_froms=True)
                    .where(Student.id.in_(selected_student_ids))
                )
            )
            if accessible_ids != set(selected_student_ids):
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="选择中包含不存在或无权导出的学生")

    def worker(task_id: str, worker_db: Session) -> dict[str, Any]:
        update_task(worker_db, task_id, 20, "正在读取学生档案")
        target = create_student_export(worker_db, filters=filters, student_ids=selected_student_ids, include_provenance=include_provenance, scope=scope, mask_sensitive=effective_mask_sensitive)
        update_task(worker_db, task_id, 90, "正在写入 XLSX 文件")
        actor = worker_db.get(User, actor_id)
        audit(worker_db, "export_xlsx", "export", target.name, actor=actor, after={"filters": filters, "selected_student_count": len(selected_student_ids or []), "selection_mode": "selected" if selected_student_ids else "all", "mask_sensitive": effective_mask_sensitive, "background_task": task_id})
        worker_db.commit()
        return {"download_url": f"/api/exports/{target.name}", "filename": target.name}

    task = submit_task(db, "xlsx_export", user, worker)
    audit(db, "queue_export_xlsx", "background_task", task.id, actor=user, after={"filters": filters, "selected_student_count": len(selected_student_ids or []), "selection_mode": "selected" if selected_student_ids else "all", "mask_sensitive": effective_mask_sensitive}, request=request)
    db.commit()
    return {"task": serialize_task(task)}


@app.post("/api/exports/preview")
def preview_student_export(
    payload: dict[str, Any] | None = Body(default=None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """Calculate an export without creating a file or recording a download."""
    _require_capability(user, "student_export", "导出")
    payload = payload or {}
    requested_filters = payload.get("filters") if isinstance(payload.get("filters"), dict) else payload
    filter_keys = ("keyword", "current_class", "school_major", "college", "school")
    filters = {
        key: str(requested_filters.get(key) or "").strip()
        for key in filter_keys
        if str(requested_filters.get(key) or "").strip()
    }
    raw_student_ids = payload.get("student_ids")
    selected_student_ids: list[int] = []
    if raw_student_ids is not None:
        if not isinstance(raw_student_ids, list):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="学生选择列表格式不正确")
        try:
            selected_student_ids = list(dict.fromkeys(int(value) for value in raw_student_ids if str(value).strip()))
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="学生选择列表格式不正确") from exc
        if len(selected_student_ids) > 10000:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="单次最多选择 10000 名学生导出")
    scope = _student_scope(db, user)
    statement = build_student_query(filters=filters, scope=scope)
    if selected_student_ids:
        statement = statement.where(Student.id.in_(selected_student_ids))
    total = int(db.scalar(select(func.count()).select_from(statement.order_by(None).subquery())) or 0)
    samples = list(db.scalars(statement.limit(8)))
    fields = [field for field in payload.get("fields", []) if field in EXPORT_FIELD_HEADERS]
    if not fields:
        fields = ["student_no", "full_name", "school", "college", "school_major", "current_class"]
    return {
        "total": total,
        "selection_mode": "selected" if selected_student_ids else "filtered",
        "filters": filters,
        "fields": [{"key": field, "label": EXPORT_FIELD_HEADERS[field]} for field in fields],
        "samples": [
            {
                "student_no": student.student_no,
                "full_name": student.full_name,
                "school": student.school,
                "college": student.college,
                "school_major": student.school_major,
                "current_class": student.current_class,
            }
            for student in samples
        ],
    }


@app.get("/api/exports/{filename}")
def download_export(filename: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if Path(filename).name != filename or not filename.endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导出文件不存在")
    target = settings.export_path / filename
    if not target.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导出文件不存在")
    return FileResponse(target, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _get_ai_conversation(db: Session, user: User, conversation_id: str | None) -> AiConversation:
    if conversation_id:
        conversation = db.get(AiConversation, conversation_id)
        if not conversation or conversation.user_id != user.id:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI 会话不存在")
        return conversation
    conversation = AiConversation(id=str(uuid4()), user_id=user.id)
    db.add(conversation)
    db.flush()
    return conversation


def _conversation_history(db: Session, conversation: AiConversation | None) -> list[dict[str, str]]:
    if not conversation:
        return []
    messages = list(
        db.scalars(
            select(AiConversationMessage)
            .where(AiConversationMessage.conversation_id == conversation.id)
            .order_by(AiConversationMessage.id.desc())
            .limit(8)
        )
    )
    history = []
    for message in reversed(messages):
        content = message.content
        if message.role == "assistant":
            state = next(
                (item.get("state") for item in (message.sources or []) if isinstance(item, dict) and item.get("type") == "tool_state" and isinstance(item.get("state"), dict)),
                None,
            )
            if state:
                content += "\n[可信工具状态] " + json.dumps(state, ensure_ascii=False, separators=(",", ":"))
        history.append({"role": message.role, "content": content})
    return history


def _ai_database_sources(filters: dict[str, str] | None, count: int | None = None) -> list[dict[str, str]]:
    descriptions = []
    for field, value in (filters or {}).items():
        if field == "keyword":
            descriptions.append(f"关键词包含“{value}”")
        elif field.startswith("exclude_"):
            label = RESPONSE_FIELD_LABELS.get(field.removeprefix("exclude_"), field.removeprefix("exclude_"))
            descriptions.append(f"{label}不包含“{value}”")
        else:
            label = RESPONSE_FIELD_LABELS.get(field, field)
            descriptions.append(f"{label}包含“{value}”")
    detail = "学生档案数据库"
    if descriptions:
        detail += "，筛选：" + "；".join(descriptions)
    if count is not None:
        detail += f"，匹配 {count} 条"
    sources = [{"type": "database", "title": "学生档案数据库", "detail": detail}]
    if count is not None:
        sources.append({"type": "method", "title": "统计口径", "detail": "按当前登录账号的数据范围、筛选条件和学生主档案实时计算"})
    return sources


def _ai_model_sources() -> list[dict[str, str]]:
    return [{"type": "model", "title": "本地 AI 模型", "detail": settings.ollama_model}]


def _related_excel_header(payload: dict[str, Any], column_index: int) -> str:
    header_rows = payload.get("header_rows") if isinstance(payload, dict) else []
    if not isinstance(header_rows, list):
        return f"第 {column_index + 1} 列"
    labels: list[str] = []
    merged_ranges = payload.get("merged_ranges") if isinstance(payload, dict) else []
    for row_index, row in enumerate(header_rows):
        value = row[column_index] if isinstance(row, list) and column_index < len(row) else ""
        if not value and isinstance(merged_ranges, list):
            merged = next(
                (
                    item for item in merged_ranges
                    if isinstance(item, dict)
                    and int(item.get("start_row") or 0) <= row_index + 1 <= int(item.get("end_row") or 0)
                    and int(item.get("start_column") or 0) <= column_index + 1 <= int(item.get("end_column") or 0)
                ),
                None,
            )
            if merged:
                start_row = int(merged.get("start_row") or 1) - 1
                start_column = int(merged.get("start_column") or 1) - 1
                source_row = header_rows[start_row] if start_row < len(header_rows) and isinstance(header_rows[start_row], list) else []
                value = source_row[start_column] if start_column < len(source_row) else ""
        text = str(value or "").strip()
        if text and text not in labels:
            labels.append(text)
    return " / ".join(labels) or f"第 {column_index + 1} 列"


def _related_value_for_user(db: Session, user: User, header: str, value: Any) -> str:
    return str(value or "").strip()


def _ai_related_info_for_student(
    db: Session,
    user: User,
    student: Student,
    question: str,
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, Any]]]:
    """Read approved cards and reviewable Excel rows for a student, read-only."""
    cards = list(
        db.scalars(
            select(StudentRelatedInfoCard)
            .where(StudentRelatedInfoCard.student_id == student.id)
            .order_by(StudentRelatedInfoCard.imported_at.desc(), StudentRelatedInfoCard.id.desc())
        )
    )
    entries: list[dict[str, Any]] = [
        {
            "key": f"card_{card.id}",
            "title": card.title,
            "payload": card.excel_payload or {},
            "document_id": card.source_document_id,
            "pending": False,
        }
        for card in cards
    ]
    # Reviewers can use a correctly matched Excel row to answer an operational
    # question before approval, but it remains visibly pending and is never
    # written to the student archive by this read-only assistant path.
    if user.role in {Role.SUPER_ADMIN, Role.ADMIN} and has_capability(user, "related_review"):
        pending_candidates = list(
            db.scalars(
                select(RelatedInfoCandidate)
                .where(
                    RelatedInfoCandidate.student_id == student.id,
                    RelatedInfoCandidate.status == CandidateStatus.PENDING,
                    RelatedInfoCandidate.content_type == "excel_card",
                )
                .order_by(RelatedInfoCandidate.created_at.desc(), RelatedInfoCandidate.id.desc())
            )
        )
        for candidate in pending_candidates:
            if not candidate.excel_payload:
                continue
            document = db.get(SourceDocument, candidate.source_document_id)
            entries.append(
                {
                    "key": f"pending_{candidate.id}",
                    "title": Path(document.original_filename).stem if document else "待审核相关资料",
                    "payload": candidate.excel_payload,
                    "document_id": candidate.source_document_id,
                    "pending": True,
                }
            )
    normalized = "".join(str(question or "").split())
    broad_query = any(term in normalized for term in ("附属表格", "附属资料", "相关资料", "原始资料", "词条"))
    scholarship_query = any(term in normalized for term in ("奖学金", "获奖", "奖项", "什么奖", "获得过", "评优", "荣誉", "三好学生", "奖励"))
    residence_query = any(term in normalized for term in ("外宿", "走读", "外住", "住宿"))
    topics: list[dict[str, Any]] = []
    if scholarship_query:
        topics.append({"label": "奖学金", "terms": ("奖学金", "获奖", "奖项", "评优", "荣誉", "三好学生", "奖励"), "found": False, "card_titles": [], "fact_keys": []})
    if residence_query:
        topics.append({"label": "外宿", "terms": ("外宿", "走读", "外住", "住宿"), "found": False, "card_titles": [], "fact_keys": []})
    if not topics and broad_query:
        topics.append({"label": "相关资料", "terms": (), "found": False, "card_titles": [], "fact_keys": []})
    facts: list[dict[str, str]] = []
    sources: list[dict[str, str]] = []
    seen_sources: set[tuple[str, str, str]] = set()
    for entry in entries:
        payload = entry["payload"]
        row = payload.get("data_row") if isinstance(payload, dict) else None
        if not isinstance(row, list):
            continue
        document = db.get(SourceDocument, entry["document_id"])
        document_text = document.original_filename if document else ""
        card_text = f"{entry['title']} {document_text}".replace(" ", "")
        matched_topics: list[dict[str, Any]] = []
        for column_index, value in enumerate(row):
            value_text = str(value or "").strip()
            if not value_text:
                continue
            header = _related_excel_header(payload, column_index)
            header_text = header.replace(" ", "")
            for topic in topics:
                topic_terms = topic["terms"]
                topic_card_match = not topic_terms or any(term in card_text for term in topic_terms)
                topic_header_match = any(term in header_text for term in topic_terms)
                if not (broad_query or topic_card_match or topic_header_match):
                    continue
                topic["found"] = True
                source_title = f"{entry['title']}（待审核）" if entry["pending"] else entry["title"]
                if source_title not in topic["card_titles"]:
                    topic["card_titles"].append(source_title)
                matched_topics.append(topic)
                # For a named topic, expose matching columns. Broad related-data
                # questions retain the prior behavior and expose the whole row.
                relevant = broad_query or topic_header_match or (topic["label"] == "奖学金" and topic_card_match and any(term in header_text for term in ("奖学金", "奖项", "三好学生", "评优", "荣誉", "奖励")))
                if not relevant:
                    continue
                field = f"related_{entry['key']}_{column_index}"
                if any(item["field"] == field for item in facts):
                    continue
                fact = {"field": field, "label": f"{entry['title']} · {header}", "value": _related_value_for_user(db, user, header, value_text)}
                facts.append(fact)
                topic["fact_keys"].append(field)
        if matched_topics:
            sheet = str(payload.get("sheet_name") or "首个可见工作表") if isinstance(payload, dict) else "首个可见工作表"
            row_number = payload.get("source_row") if isinstance(payload, dict) else None
            source_key = (entry["title"], sheet, str(row_number or ""))
            if source_key not in seen_sources:
                seen_sources.add(source_key)
                detail = f"工作表 {sheet}"
                if row_number:
                    detail += f" / 第 {row_number} 行"
                if document:
                    detail = f"{document.original_filename} · {detail}"
                if entry["pending"]:
                    detail += " · 待审核"
                sources.append({"type": "related_info_pending" if entry["pending"] else "related_info", "title": entry["title"], "detail": detail})
    return facts, sources, topics


def _normalize_ai_name(value: str | None) -> str:
    return "".join(character for character in str(value or "") if "\u4e00" <= character <= "\u9fff")


def _find_ai_fuzzy_name_match(
    db: Session,
    filters: dict[str, str],
    scope: list[dict[str, str]],
) -> tuple[list[Student], dict[str, Any] | None]:
    """Resolve one clearly misspelled Chinese name without relaxing account scope."""
    raw_name = next((filters.get(field) for field in ("full_name", "keyword") if filters.get(field)), None)
    requested_name = _normalize_ai_name(raw_name)
    if len(requested_name) < 3 or len(requested_name) > 4:
        return [], None
    base_filters = {field: value for field, value in filters.items() if field not in {"full_name", "keyword"}}
    candidates = list(
        db.scalars(
            build_student_query(filters=base_filters, scope=scope)
            .order_by(None)
            .where(Student.full_name.is_not(None), Student.full_name != "")
            .limit(500)
        )
    )
    matches: list[tuple[float, Student]] = []
    for student in candidates:
        candidate_name = _normalize_ai_name(student.full_name)
        if len(candidate_name) != len(requested_name) or not candidate_name or candidate_name[0] != requested_name[0]:
            continue
        same_positions = sum(left == right for left, right in zip(requested_name, candidate_name))
        # A 3-character Chinese name needs two matching positions, and a
        # 4-character name needs three. This permits a single typo but avoids
        # broad name guessing across the archive.
        if same_positions < len(requested_name) - 1:
            continue
        matches.append((same_positions / len(requested_name), student))
    matches.sort(key=lambda item: (-item[0], item[1].student_no))
    if not matches or (len(matches) > 1 and matches[0][0] == matches[1][0]):
        return [], None
    score, student = matches[0]
    return [student], {"query": requested_name, "matched_name": student.full_name, "similarity": round(score, 2)}


def _award_summary_from_details(details: list[dict[str, Any]]) -> tuple[list[str], list[str]]:
    """Convert verified award cells to natural language and retain every award fact."""
    awards: list[str] = []
    required_terms: list[str] = []
    named_scholarships = ("国家奖学金", "励志奖学金", "省政府奖学金", "创新创业奖学金")
    for detail in details:
        header = str(detail.get("label") or "").split(" · ", 1)[-1]
        value = str(detail.get("value") or "").strip()
        if not value:
            continue
        if "拟评奖项" in header or "三好学生" in header:
            award = value
            required_terms.append(value)
        else:
            scholarship = next((item for item in named_scholarships if item in header), None)
            if not scholarship:
                continue
            if value in {"√", "是", "有", "1", "true", "True"}:
                award = scholarship
                required_terms.append(scholarship)
            else:
                award = scholarship if value in scholarship else f"{scholarship}{value}"
                required_terms.extend([scholarship, value])
        if award not in awards:
            awards.append(award)
    return awards, list(dict.fromkeys(term for term in required_terms if term))


def _with_ai_tool_state(response: dict[str, Any], plan: dict[str, Any]) -> dict[str, Any]:
    state = {
        "intent": response.get("intent") or plan.get("intent"),
        "filters": response.get("_resolved_filters") or plan.get("filters") or {},
        "fields": plan.get("fields") or [],
    }
    if plan.get("aggregation"):
        state["aggregation"] = plan["aggregation"]
    if plan.get("top_group_by"):
        state["top_group_by"] = plan["top_group_by"]
    if plan.get("filename_stem"):
        state["filename_stem"] = plan["filename_stem"]
    return {**response, "_tool_state": state}


def _save_ai_conversation_turn(
    db: Session,
    conversation: AiConversation | None,
    user: User,
    question: str,
    response: dict[str, Any],
    request: Request,
) -> dict[str, Any]:
    if not conversation:
        return {key: value for key, value in response.items() if not key.startswith("_")}
    tool_state = response.get("_tool_state")
    stored_sources = list(response.get("sources") or [])
    if isinstance(tool_state, dict):
        stored_sources.append({"type": "tool_state", "state": tool_state})
    public_response = {key: value for key, value in response.items() if not key.startswith("_")}
    db.add(AiConversationMessage(conversation_id=conversation.id, role="user", content=question, intent="question"))
    db.add(
        AiConversationMessage(
            conversation_id=conversation.id,
            role="assistant",
            content=_stream_text(public_response)[:4000],
            sources=stored_sources,
            intent=public_response.get("intent"),
        )
    )
    conversation.updated_at = utcnow()
    audit(
        db,
        "ai_conversation_turn",
        "ai_conversation",
        conversation.id,
        actor=user,
        after={
            "question": question[:1000],
            "reply": _stream_text(public_response)[:4000],
            "intent": public_response.get("intent"),
            "sources": stored_sources,
        },
        request=request,
    )
    db.commit()
    return {**public_response, "conversation_id": conversation.id}


def _create_ai_pending_action(
    db: Session,
    user: User,
    conversation: AiConversation | None,
    action_type: str,
    payload: dict[str, Any],
) -> AiPendingAction:
    action = AiPendingAction(
        id=str(uuid4()),
        user_id=user.id,
        conversation_id=conversation.id if conversation else None,
        action_type=action_type,
        payload=payload,
        expires_at=utcnow() + timedelta(minutes=10),
    )
    db.add(action)
    db.flush()
    return action


def _append_ai_action_result(db: Session, action: AiPendingAction, response: dict[str, Any]) -> None:
    if not action.conversation_id:
        return
    conversation = db.get(AiConversation, action.conversation_id)
    if not conversation:
        return
    db.add(
        AiConversationMessage(
            conversation_id=conversation.id,
            role="assistant",
            content=_stream_text(response)[:4000],
            sources=response.get("sources") or [],
            intent=response.get("intent"),
        )
    )
    conversation.updated_at = utcnow()


def _execute_ai_aggregation(plan: dict[str, Any], question: str, request: Request, db: Session, user: User) -> dict[str, Any]:
    aggregation = plan["aggregation"]
    operation = aggregation["operation"]
    field = aggregation.get("field")
    value = aggregation.get("value")
    filters = dict(plan.get("filters") or {})
    scope = _student_scope(db, user)
    if operation == "count" and field and value:
        filters.setdefault(field, value)
    column = getattr(Student, field) if field else None

    if operation == "count":
        count = _count_students(db, filters, scope)
        if not filters:
            reply = f"当前学生档案共有 {count} 人。"
        elif len(filters) == 1:
            filter_field, filter_value = next(iter(filters.items()))
            if filter_field.startswith("exclude_"):
                label = RESPONSE_FIELD_LABELS[filter_field.removeprefix("exclude_")]
                reply = f"{label}不包含{filter_value}的学生共有 {count} 人。"
            else:
                reply = f"{filter_value}的{RESPONSE_FIELD_LABELS[filter_field]}共有 {count} 人。"
        else:
            condition_text = "、".join(
                f"{RESPONSE_FIELD_LABELS[field.removeprefix('exclude_')]}不包含{value}"
                if field.startswith("exclude_")
                else f"{RESPONSE_FIELD_LABELS[field]}为{value}"
                for field, value in filters.items()
            )
            reply = f"{condition_text}的学生共有 {count} 人。"
        audit(db, "ai_aggregate", "student_aggregate", "assistant", actor=user, after={"question": question, "aggregation": aggregation, "filters": filters, "count": count}, request=request)
        db.commit()
        return {"reply": reply, "intent": "aggregate", "results": [], "sources": _ai_database_sources(filters, count)}

    if column is None:
        return {"reply": "请说明需要按哪个字段统计分布，例如学校专业、所在班级或所属学院。", "intent": "answer", "results": [], "sources": _ai_model_sources()}
    filtered_students = build_student_query(filters=filters, scope=scope).order_by(None).subquery()
    grouped_column = filtered_students.c[field]
    rows = list(
        db.execute(
            select(grouped_column, func.count())
            .where(grouped_column.is_not(None), grouped_column != "")
            .group_by(grouped_column)
            .order_by(func.count().desc(), grouped_column.asc())
            .limit(30)
        )
    )
    if not rows:
        reply = f"目前没有可用于统计的{AGGREGATE_FIELD_LABELS[field]}数据。"
    else:
        reply = f"{AGGREGATE_FIELD_LABELS[field]}人数分布：\n" + "\n".join(f"{item_value}: {item_count} 人" for item_value, item_count in rows)
    audit(db, "ai_aggregate", "student_aggregate", "assistant", actor=user, after={"question": question, "aggregation": aggregation, "filters": filters, "groups": len(rows)}, request=request)
    db.commit()
    return {"reply": reply, "intent": "aggregate", "results": [], "sources": _ai_database_sources(filters, len(rows))}


def _execute_ai_top_group_search(plan: dict[str, Any], question: str, request: Request, db: Session, user: User) -> dict[str, Any]:
    field = str(plan.get("top_group_by") or "")
    if field not in AGGREGATE_FIELD_LABELS:
        return {"reply": "请先说明需要按哪个维度统计人数。", "intent": "answer", "results": [], "sources": _ai_model_sources()}
    scope = _student_scope(db, user)
    base_filters = dict(plan.get("filters") or {})
    students = build_student_query(filters=base_filters, scope=scope).order_by(None).subquery()
    grouped_column = students.c[field]
    top_group = db.execute(
        select(grouped_column, func.count())
        .where(grouped_column.is_not(None), grouped_column != "")
        .group_by(grouped_column)
        .order_by(func.count().desc(), grouped_column.asc())
        .limit(1)
    ).first()
    if not top_group:
        return {"reply": f"目前没有可用于统计的{AGGREGATE_FIELD_LABELS[field]}数据。", "intent": "answer", "results": [], "sources": _ai_database_sources({}, 0)}
    group_value, total = top_group
    resolved_filters = {**base_filters, field: str(group_value)}
    records = list_students(db, filters=resolved_filters, limit=50, scope=scope)
    display_fields = plan.get("fields") or ["student_no", "full_name"]
    results = []
    for item in records:
        serialized = _serialize_student_for_user(db, user, item)
        results.append(
            {
                "id": item.id,
                "student_no": item.student_no,
                "full_name": item.full_name,
                "details": [
                    {"field": detail_field, "label": RESPONSE_FIELD_LABELS[detail_field], "value": serialized.get(detail_field) or "未登记"}
                    for detail_field in display_fields
                ],
            }
        )
    reply = f"{AGGREGATE_FIELD_LABELS[field]}人数最多的是{group_value}，共有 {total} 人。已列出 {len(records)} 条学生记录。"
    audit(db, "ai_top_group_search", "student_query", "assistant", actor=user, after={"question": question, "field": field, "value": group_value, "total": total}, request=request)
    db.commit()
    return {
        "reply": reply,
        "intent": "search",
        "results": results,
        "sources": _ai_database_sources(resolved_filters, total),
        "_resolved_filters": resolved_filters,
    }


def _execute_ai_question(
    question: str,
    request: Request,
    db: Session,
    user: User,
    conversation: AiConversation | None = None,
) -> dict[str, Any]:
    plan = plan_assistant_question(question, _conversation_history(db, conversation))
    controls = get_controls(db)
    if plan["intent"] in {"answer", "unsupported"}:
        return _save_ai_conversation_turn(db, conversation, user, question, _with_ai_tool_state({"reply": plan["reply"], "intent": plan["intent"], "results": [], "sources": _ai_model_sources()}, plan), request)
    if plan["intent"] == "aggregate":
        return _save_ai_conversation_turn(db, conversation, user, question, _with_ai_tool_state(_execute_ai_aggregation(plan, question, request, db, user), plan), request)
    if plan.get("top_group_by"):
        return _save_ai_conversation_turn(db, conversation, user, question, _with_ai_tool_state(_execute_ai_top_group_search(plan, question, request, db, user), plan), request)
    filters = plan["filters"]
    scope = _student_scope(db, user)
    if plan["intent"] == "bulk_update":
        return _save_ai_conversation_turn(db, conversation, user, question, _with_ai_tool_state({"reply": "AI 数据助手处于只读模式，不能修改学生档案或数据库。请使用平台编辑功能完成修改。", "intent": "answer", "results": [], "sources": _ai_model_sources()}, plan), request)
    if plan["intent"] == "export" and not controls["ai_operations_enabled"]:
        return _save_ai_conversation_turn(db, conversation, user, question, _with_ai_tool_state({"reply": "系统设置已关闭 AI 发起的数据操作；你仍可使用 AI 查询和统计。", "intent": "answer", "results": [], "sources": _ai_model_sources()}, plan), request)
    if plan["intent"] == "export":
        _require_capability(user, "student_export", "导出")
        requested_fields = plan.get("fields", [])
        count = _count_students(db, filters, scope)
        if not controls["ai_export_confirmation_required"]:
            target = create_student_export(db, filters=filters, fields=requested_fields or None, scope=scope, mask_sensitive=False, filename_stem=plan.get("filename_stem"))
            response = {"reply": f"已生成 {count} 条学生记录的 XLSX 文件：{target.name}。", "intent": "export", "download_url": f"/api/exports/{target.name}", "results": [], "sources": _ai_database_sources(filters, count)}
            audit(db, "ai_export_xlsx", "export", target.name, actor=user, after={"question": question, "filters": filters, "fields": requested_fields, "filename_stem": plan.get("filename_stem"), "filename": target.name, "count": count, "mask_sensitive": False}, request=request)
            return _save_ai_conversation_turn(db, conversation, user, question, _with_ai_tool_state(response, plan), request)
        filename_stem = plan.get("filename_stem")
        action = _create_ai_pending_action(db, user, conversation, "export", {"filters": filters, "fields": requested_fields, "filename_stem": filename_stem, "question": question})
        response = {
            "reply": f"将导出 {count} 条学生记录{f'，文件名为 {filename_stem}.xlsx' if filename_stem else ''}。请确认后生成 XLSX 文件。",
            "intent": "export_confirmation",
            "results": [],
            "confirmation": {"action_id": action.id, "action_type": "export", "label": "确认导出 XLSX"},
            "sources": _ai_database_sources(filters, count),
        }
        audit(db, "ai_prepare_export", "ai_pending_action", action.id, actor=user, after={"question": question, "filters": filters, "fields": requested_fields, "filename_stem": filename_stem, "count": count}, request=request)
        return _save_ai_conversation_turn(db, conversation, user, question, _with_ai_tool_state(response, plan), request)
    records = list_students(db, filters=filters, limit=50, scope=scope)
    fuzzy_match: dict[str, Any] | None = None
    if not records:
        records, fuzzy_match = _find_ai_fuzzy_name_match(db, filters, scope)
    requested_fields = plan.get("fields", [])
    display_fields = requested_fields or ["school", "college", "current_class"]
    results = []
    related_sources: list[dict[str, str]] = []
    related_result_count = 0
    related_topics_for_response: list[dict[str, Any]] = []
    for item in records:
        details = []
        for field in display_fields:
            value = _serialize_student_for_user(db, user, item).get(field)
            if hasattr(value, "isoformat"):
                value = value.isoformat()
            details.append({"field": field, "label": RESPONSE_FIELD_LABELS[field], "value": value or "未登记"})
        related_facts, card_sources, related_topics = _ai_related_info_for_student(db, user, item, question)
        if related_facts:
            related_result_count += 1
            details.extend(related_facts)
            related_sources.extend(card_sources)
        if len(records) == 1:
            related_topics_for_response = related_topics
        results.append({"id": item.id, "student_no": item.student_no, "full_name": item.full_name, "details": details})
    fuzzy_note = ""
    if fuzzy_match:
        fuzzy_note = f"未检索到姓名“{fuzzy_match['query']}”的精确记录，已按近似姓名匹配到“{fuzzy_match['matched_name']}”。"
    required_response_terms: list[str] = ["近似", str(fuzzy_match["matched_name"])] if fuzzy_match else []
    if not records:
        reply = "没有找到符合条件的学生记录。"
    elif requested_fields and len(records) == 1:
        detail_text = "，".join(f"{detail['label']}是{detail['value']}" for detail in results[0]["details"])
        reply = f"{fuzzy_note}{records[0].full_name}的{detail_text}。"
    elif _is_related_info_query(question) and len(records) == 1:
        related_details = [detail for detail in results[0]["details"] if str(detail.get("field") or "").startswith("related_")]
        topic_parts: list[str] = [fuzzy_note] if fuzzy_note else []
        for topic in related_topics_for_response:
            topic_facts = [detail for detail in related_details if detail.get("field") in topic.get("fact_keys", [])]
            if topic["label"] == "奖学金":
                if topic["found"] and topic_facts:
                    awards, award_terms = _award_summary_from_details(topic_facts)
                    required_response_terms.extend(award_terms)
                    if awards:
                        topic_parts.append(f"{records[0].full_name}拿过奖学金，记录显示：{'、'.join(awards)}。")
                    else:
                        detail_text = "；".join(f"{detail['label'].split(' · ', 1)[-1]}：{detail['value']}" for detail in topic_facts)
                        topic_parts.append(f"{records[0].full_name}拿过奖学金，相关记录为：{detail_text}。")
                else:
                    topic_parts.append(f"未检索到{records[0].full_name}的奖学金相关信息。")
            elif topic["label"] == "外宿":
                if topic["found"]:
                    title_text = "、".join(topic["card_titles"])
                    topic_parts.append(f"已在{records[0].full_name}的外宿相关表格中检索到信息，来源：{title_text}。")
                else:
                    topic_parts.append(f"未检索到{records[0].full_name}的相关外宿信息。")
            elif topic["found"] and topic_facts:
                detail_text = "；".join(f"{detail['label']}：{detail['value']}" for detail in topic_facts)
                topic_parts.append(f"{records[0].full_name}的相关资料显示：{detail_text}。")
            else:
                topic_parts.append(f"已找到{records[0].full_name}，但其已审核的相关资料中没有匹配到对应信息。")
        if topic_parts:
            reply = " ".join(topic_parts)
        elif related_details:
            detail_text = "；".join(f"{detail['label']}：{detail['value']}" for detail in related_details)
            reply = f"{records[0].full_name}的相关资料显示：{detail_text}。"
        else:
            reply = f"已找到{records[0].full_name}，但其已审核的相关资料中没有匹配到“{question}”对应的信息。"
    elif requested_fields:
        reply = f"找到 {len(records)} 条记录，已列出所需字段。"
    elif fuzzy_match and len(records) == 1:
        student = records[0]
        reply = f"{fuzzy_note}{student.full_name}的学号是{student.student_no}，所属学校为{student.school or '未登记'}，所属学院为{student.college or '未登记'}，所在班级为{student.current_class or '未登记'}。"
    else:
        reply = f"{plan['reply']} 共找到 {len(records)} 条记录。"
    response_facts = {
        "查询类型": "学生相关信息" if _is_related_info_query(question) else "学生档案查询",
        "匹配人数": len(records),
        "学生": results[:20],
        "相关资料主题": related_topics_for_response,
        "近似姓名匹配": fuzzy_match,
        "必须保留的事实": list(dict.fromkeys(required_response_terms)),
    }
    reply = express_assistant_answer(question, response_facts, reply, required_terms=required_response_terms)
    resolved_filters = dict(filters)
    fuzzy_sources: list[dict[str, str]] = []
    if fuzzy_match:
        resolved_filters.pop("keyword", None)
        resolved_filters.pop("full_name", None)
        resolved_filters["full_name"] = str(fuzzy_match["matched_name"])
        fuzzy_sources.append({"type": "fuzzy_match", "title": "近似姓名匹配", "detail": f"未找到“{fuzzy_match['query']}”的精确记录，按单字差异匹配到“{fuzzy_match['matched_name']}”"})
    audit(db, "ai_search", "student_query", "assistant", actor=user, after={"question": question, "filters": filters, "resolved_filters": resolved_filters, "fields": requested_fields, "count": len(records), "fuzzy_match": fuzzy_match}, request=request)
    db.commit()
    return _save_ai_conversation_turn(db, conversation, user, question, _with_ai_tool_state({
        "reply": reply,
        "intent": "search",
        "results": results,
        "sources": _ai_database_sources(resolved_filters, len(records)) + fuzzy_sources + related_sources,
        "_resolved_filters": resolved_filters,
    }, plan), request)


def _stream_text(response: dict[str, Any]) -> str:
    # Results are machine-readable evidence for the model and API, not a second
    # transcript to append beneath the assistant's natural-language response.
    return str(response.get("reply") or "")


def _sse_event(event: str, payload: dict[str, Any]) -> str:
    return f"event: {event}\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"


def _stream_ai_response(response: dict[str, Any]):
    for character in _stream_text(response):
        yield _sse_event("delta", {"text": character})
        time.sleep(0.012)
    yield _sse_event("done", response)


@app.post("/api/ai/chat")
def ask_ai(question: AiQuestion, request: Request, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict[str, Any]:
    require_csrf(request)
    conversation = _get_ai_conversation(db, user, str(question.conversation_id) if question.conversation_id else None)
    started = time.perf_counter()
    response = _execute_ai_question(question.question, request, db, user, conversation)
    duration_ms = int((time.perf_counter() - started) * 1000)
    message = db.scalar(select(AiConversationMessage).where(AiConversationMessage.conversation_id == conversation.id, AiConversationMessage.role == "assistant").order_by(AiConversationMessage.id.desc()).limit(1))
    if message:
        message.model_name = settings.ollama_model
        message.duration_ms = duration_ms
        db.commit()
    response["ai_metrics"] = {"model": settings.ollama_model, "duration_ms": duration_ms}
    return response


@app.post("/api/ai/chat/stream")
def ask_ai_stream(question: AiQuestion, request: Request, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    require_csrf(request)
    conversation = _get_ai_conversation(db, user, str(question.conversation_id) if question.conversation_id else None)
    started = time.perf_counter()
    response = _execute_ai_question(question.question, request, db, user, conversation)
    duration_ms = int((time.perf_counter() - started) * 1000)
    message = db.scalar(select(AiConversationMessage).where(AiConversationMessage.conversation_id == conversation.id, AiConversationMessage.role == "assistant").order_by(AiConversationMessage.id.desc()).limit(1))
    if message:
        message.model_name = settings.ollama_model
        message.duration_ms = duration_ms
        db.commit()
    response["ai_metrics"] = {"model": settings.ollama_model, "duration_ms": duration_ms}
    return StreamingResponse(_stream_ai_response(response), media_type="text/event-stream", headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.post("/api/ai/actions/{action_id}/confirm")
def confirm_ai_action(
    action_id: str,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    require_csrf(request)
    action = db.get(AiPendingAction, action_id)
    if not action or action.user_id != user.id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="待确认操作不存在")
    if action.status != "pending":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该操作已处理")
    expires_at = action.expires_at
    expires_at = as_china_time(expires_at)
    if expires_at < utcnow():
        action.status = "expired"
        db.commit()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="确认已过期，请重新发起操作")
    payload = action.payload or {}
    filters = payload.get("filters") if isinstance(payload.get("filters"), dict) else {}
    scope = _student_scope(db, user)
    if action.action_type == "export":
        _require_capability(user, "student_export", "导出")
        fields = payload.get("fields") if isinstance(payload.get("fields"), list) else []
        filename_stem = str(payload.get("filename_stem") or "").strip() or None
        target = create_student_export(db, filters=filters, fields=fields or None, scope=scope, mask_sensitive=False, filename_stem=filename_stem)
        count = _count_students(db, filters, scope)
        response = {
            "reply": f"已生成 {count} 条学生记录的 XLSX 文件：{target.name}。",
            "intent": "export",
            "download_url": f"/api/exports/{target.name}",
            "results": [],
            "sources": _ai_database_sources(filters, count),
        }
        audit(db, "ai_export_xlsx", "export", target.name, actor=user, after={"action_id": action.id, "question": payload.get("question"), "filters": filters, "fields": fields, "filename_stem": filename_stem, "filename": target.name, "count": count, "mask_sensitive": False}, request=request)
    elif action.action_type == "bulk_update":
        action.status = "rejected"
        audit(db, "reject_ai_bulk_update", "ai_pending_action", action.id, actor=user, after={"reason": "AI is read-only"}, request=request)
        db.commit()
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="AI 数据助手仅允许读取数据，不能修改数据库")
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="不支持的待确认操作")
    action.status = "confirmed"
    action.confirmed_at = utcnow()
    _append_ai_action_result(db, action, response)
    db.commit()
    return response


@app.get("/api/ai/conversations/{conversation_id}")
def get_ai_conversation(conversation_id: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict[str, Any]:
    conversation = _get_ai_conversation(db, user, conversation_id)
    messages = list(
        db.scalars(
            select(AiConversationMessage)
            .where(AiConversationMessage.conversation_id == conversation.id)
            .order_by(AiConversationMessage.id.asc())
            .limit(100)
        )
    )
    return {
        "conversation_id": conversation.id,
        "messages": [{"role": message.role, "content": message.content, "sources": message.sources or [], "intent": message.intent, "model_name": message.model_name, "duration_ms": message.duration_ms} for message in messages],
    }


@app.get("/api/ai/admin/conversations")
def admin_ai_conversations(
    per_user_limit: int = 10,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> list[dict[str, Any]]:
    _require_capability(user, "audit_view", "审计查看")
    safe_per_user_limit = min(max(per_user_limit, 1), 50)
    ranked_conversations = (
        select(
            AiConversation.id.label("conversation_id"),
            func.row_number()
            .over(partition_by=AiConversation.user_id, order_by=(AiConversation.updated_at.desc(), AiConversation.id.desc()))
            .label("row_number"),
        )
        .subquery()
    )
    conversation_statement = (
        select(AiConversation)
        .join(ranked_conversations, ranked_conversations.c.conversation_id == AiConversation.id)
        .where(ranked_conversations.c.row_number <= safe_per_user_limit)
    )
    if user.role == Role.ADMIN:
        conversation_statement = conversation_statement.join(User, User.id == AiConversation.user_id).where(User.role == Role.TEACHER)
    conversations = list(db.scalars(conversation_statement.order_by(AiConversation.updated_at.desc(), AiConversation.id.desc())))
    user_ids = {conversation.user_id for conversation in conversations}
    users = {item.id: item for item in db.scalars(select(User).where(User.id.in_(user_ids)))} if user_ids else {}
    result = []
    for conversation in conversations:
        messages = list(
            db.scalars(
                select(AiConversationMessage)
                .where(AiConversationMessage.conversation_id == conversation.id)
                .order_by(AiConversationMessage.id.desc())
                .limit(12)
            )
        )
        question = next((message.content for message in messages if message.role == "user"), "")
        reply = next((message.content for message in messages if message.role == "assistant"), "")
        owner = users.get(conversation.user_id)
        result.append(
            {
                "id": conversation.id,
                "user": (owner.display_name or owner.username) if owner else "已删除用户",
                "username": owner.username if owner else "",
                "updated_at": conversation.updated_at,
                "question": question[:300],
                "reply": reply[:500],
                "message_count": len(messages),
            }
        )
    return result


@app.get("/api/ai/admin/conversations/{conversation_id}")
def admin_ai_conversation_detail(
    conversation_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    _require_capability(user, "audit_view", "审计查看")
    conversation = db.get(AiConversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI 会话不存在")
    owner = db.get(User, conversation.user_id)
    if user.role == Role.ADMIN and (not owner or owner.role != Role.TEACHER):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI 会话不存在")
    messages = list(
        db.scalars(
            select(AiConversationMessage)
            .where(AiConversationMessage.conversation_id == conversation.id)
            .order_by(AiConversationMessage.id.asc())
        )
    )
    return {
        "id": conversation.id,
        "user": (owner.display_name or owner.username) if owner else "已删除用户",
        "messages": [{"role": message.role, "content": message.content, "sources": message.sources or [], "intent": message.intent, "model_name": message.model_name, "duration_ms": message.duration_ms, "created_at": message.created_at} for message in messages],
    }


@app.delete("/api/ai/conversations/{conversation_id}")
def clear_ai_conversation(conversation_id: str, request: Request, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict[str, bool]:
    require_csrf(request)
    conversation = _get_ai_conversation(db, user, conversation_id)
    db.execute(delete(AiConversationMessage).where(AiConversationMessage.conversation_id == conversation.id))
    db.delete(conversation)
    audit(db, "clear_ai_conversation", "ai_conversation", conversation.id, actor=user, request=request)
    db.commit()
    return {"ok": True}


def _append_ai_suggestion(suggestions: list[dict[str, str]], key: str, question: str) -> None:
    if not question or any(item["question"] == question for item in suggestions):
        return
    suggestions.append({"id": key, "label": question, "question": question})


@app.get("/api/ai/suggestions")
def ai_suggestions(db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> dict[str, Any]:
    """Suggest only questions whose named data is already in the caller's scope."""
    scope = _student_scope(db, user)
    visible_students = list(
        db.scalars(
            build_student_query(scope=scope)
            .order_by(None)
            .order_by(Student.updated_at.desc(), Student.student_no.asc())
            .limit(300)
        )
    )
    if not visible_students:
        return {
            "suggestions": [
                {"id": "visible-count", "label": "当前学生档案共有多少人？", "question": "当前学生档案共有多少人？"},
                {"id": "ai-help", "label": "你可以帮我做什么？", "question": "你可以帮我做什么？"},
            ]
        }

    recent_questions = "\n".join(
        str(item)
        for item in db.scalars(
            select(AiConversationMessage.content)
            .join(AiConversation, AiConversationMessage.conversation_id == AiConversation.id)
            .where(AiConversation.user_id == user.id, AiConversationMessage.role == "user")
            .order_by(AiConversationMessage.created_at.desc(), AiConversationMessage.id.desc())
            .limit(60)
        )
    )
    visible_ids = {student.id for student in visible_students}
    related_student_ids = set(
        db.scalars(
            select(StudentRelatedInfoCard.student_id)
            .where(StudentRelatedInfoCard.student_id.in_(visible_ids))
            .distinct()
        )
    ) if visible_ids else set()

    preferred: list[Student] = []

    def add_student(student: Student) -> None:
        if student.full_name and student.id not in {item.id for item in preferred}:
            preferred.append(student)

    # Use only this account's own past questions to make the prompts feel familiar.
    for student in visible_students:
        if len(student.full_name or "") >= 2 and student.full_name in recent_questions:
            add_student(student)
    for student in visible_students:
        if student.id in related_student_ids:
            add_student(student)
    for student in visible_students:
        add_student(student)

    suggestions: list[dict[str, str]] = []
    featured = preferred[0]
    _append_ai_suggestion(suggestions, f"student-profile-{featured.id}", f"查看{featured.full_name}同学的完整档案")
    if featured.id in related_student_ids:
        _append_ai_suggestion(suggestions, f"student-awards-{featured.id}", f"{featured.full_name}同学获得过什么奖？")
    _append_ai_suggestion(suggestions, f"student-export-{featured.id}", f"导出{featured.full_name}同学的所有信息")

    major = next((student.school_major for student in preferred if student.school_major), None)
    if major:
        _append_ai_suggestion(suggestions, "major-count", f"{major}专业有多少人？")
        _append_ai_suggestion(suggestions, "major-export", f"导出{major}专业学生名单")
    current_class = next((student.current_class for student in preferred if student.current_class), None)
    if current_class:
        _append_ai_suggestion(suggestions, "class-roster", f"查看{current_class}班的学生名单")
    _append_ai_suggestion(suggestions, "visible-count", "当前学生档案共有多少人？")
    return {"suggestions": suggestions[:6]}


@app.get("/api/ai/status")
def ai_status(user: User = Depends(get_current_user)) -> dict[str, str | bool]:
    return get_ai_health()


def _serialize_backup(backup: SystemBackup) -> dict[str, Any]:
    return {
        "id": backup.id,
        "file_name": backup.file_name,
        "size_bytes": backup.size_bytes,
        "database_dialect": backup.database_dialect,
        "status": backup.status,
        "error_message": backup.error_message,
        "checksum": backup.checksum,
        "validation_status": backup.validation_status,
        "validated_at": backup.validated_at,
        "storage_files": len((backup.manifest or {}).get("storage_files") or []),
        "offsite_status": (backup.manifest or {}).get("offsite", {}).get("status") if isinstance((backup.manifest or {}).get("offsite"), dict) else None,
        "created_by_id": backup.created_by_id,
        "created_at": backup.created_at,
    }


def _authorize_system_update(db: Session, user: User, payload: dict[str, Any]) -> User:
    if user.role not in {Role.SUPER_ADMIN, Role.ADMIN}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="只有管理员可以发起系统更新")
    if str(payload.get("confirmation_phrase") or "").strip() != "确认更新系统":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请输入“确认更新系统”后继续")
    username = str(payload.get("super_admin_username") or "").strip()
    password = str(payload.get("super_admin_password") or "")
    approver = db.scalar(select(User).where(User.username == username, User.role == Role.SUPER_ADMIN, User.is_active.is_(True)))
    if not approver or not verify_password(password, approver.password_hash):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="超级管理员账号或密码不正确")
    locked_until = as_china_time(approver.locked_until)
    if locked_until and locked_until > utcnow():
        raise HTTPException(status_code=status.HTTP_423_LOCKED, detail="该超级管理员账号当前已锁定")
    return approver


def _ensure_update_idle() -> None:
    current = get_update_status()
    if current.get("state") in {"downloading", "validating", "backing_up", "applying", "installing", "restarting", "rolling_back"}:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="已有系统更新任务正在执行，请等待完成或回滚")


def _schedule_update(
    db: Session,
    user: User,
    approver: User,
    request: Request,
    *,
    release: dict[str, Any] | None = None,
    offline_package: Path | None = None,
) -> dict[str, Any]:
    _ensure_update_idle()
    config = get_update_configuration(db, include_token=True)
    backup = create_database_backup(db, user)
    if backup.status != "completed":
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=backup.error_message or "更新前数据库备份失败")
    job_path, job = create_update_job(release, requested_by_id=user.id, offline_package=offline_package)
    write_update_status(
        {
            "state": "queued",
            "message": "更新已排队，正在启动独立更新器",
            "progress": 1,
            "job_id": job["job_id"],
            "requested_by": user.id,
            "approved_by": approver.id,
            "current_version": APP_RELEASE,
            "target_version": (release or {}).get("tag_name") if release else "离线更新包",
            "backup": backup.file_name,
        }
    )
    audit(
        db,
        "queue_system_update",
        "system_update",
        job["job_id"],
        actor=user,
        after={
            "source": job["source"],
            "current_version": APP_RELEASE,
            "target_version": (release or {}).get("tag_name") if release else "offline",
            "backup": backup.file_name,
            "approved_by_id": approver.id,
        },
        request=request,
    )
    db.commit()
    try:
        launch_update_runner(job_path, config.get("github_token"))
    except Exception as exc:
        write_update_status({"state": "failed", "message": "无法启动独立更新器", "progress": 100, "error": str(exc)[:1000], "job_id": job["job_id"]})
        audit(db, "fail_system_update_runner", "system_update", job["job_id"], actor=user, after={"error": str(exc)[:1000]}, request=request)
        db.commit()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"无法启动更新器: {exc}") from exc
    return {"ok": True, "job_id": job["job_id"], "backup": backup.file_name, "message": "更新器已启动，服务将在下载和校验完成后自动重启"}


@app.get("/api/system/updates")
def system_update_status(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    return {
        "current_version": APP_RELEASE,
        "configuration": get_update_configuration(db),
        "status": get_update_status(),
        "may_configure": user.role == Role.SUPER_ADMIN,
        "may_execute": user.role in {Role.SUPER_ADMIN, Role.ADMIN},
    }


@app.post("/api/system/updates/check")
def check_system_update(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
    background: bool = False,
) -> dict[str, Any]:
    require_csrf(request)
    result = check_for_update(db)
    if not background:
        audit(
            db,
            "check_system_update",
            "system_update",
            result.get("release", {}).get("tag_name") if isinstance(result.get("release"), dict) else "none",
            actor=user,
            after={"configured": result.get("configured"), "repository": result.get("repository"), "release": result.get("release", {}).get("tag_name") if isinstance(result.get("release"), dict) else None},
            request=request,
        )
        db.commit()
    return result


@app.get("/api/system/updates/config")
def get_system_update_configuration(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, Any]:
    return get_update_configuration(db)


@app.put("/api/system/updates/config")
def set_system_update_configuration(
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    current_password = str(payload.get("current_password") or "")
    if not verify_password(current_password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="当前超级管理员密码不正确")
    before = get_update_configuration(db)
    configuration = save_update_configuration(
        db,
        str(payload.get("repository") or ""),
        str(payload.get("channel") or "stable"),
        str(payload.get("github_token") or "") or None,
    )
    audit(
        db,
        "configure_system_update",
        "system_update",
        "configuration",
        actor=user,
        before=before,
        after=configuration | {"github_token_changed": bool(str(payload.get("github_token") or "").strip())},
        request=request,
    )
    db.commit()
    return configuration


@app.post("/api/system/updates/start")
def start_system_update(
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    approver = _authorize_system_update(db, user, payload)
    result = check_for_update(db)
    release = result.get("release") if isinstance(result.get("release"), dict) else None
    requested_tag = str(payload.get("tag_name") or "").strip()
    if not release or not release.get("update_ready"):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="没有可用的受控更新包，请先检查 GitHub Release")
    if requested_tag and requested_tag != release.get("tag_name"):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="可用版本已变化，请重新检查更新")
    if not release.get("is_newer"):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="当前已是最新版本，或目标版本不比当前版本新")
    return _schedule_update(db, user, approver, request, release=release)


@app.post("/api/system/updates/offline")
async def start_offline_system_update(
    request: Request,
    file: UploadFile = File(...),
    confirmation_phrase: str = Form(...),
    super_admin_username: str = Form(...),
    super_admin_password: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    approver = _authorize_system_update(
        db,
        user,
        {
            "confirmation_phrase": confirmation_phrase,
            "super_admin_username": super_admin_username,
            "super_admin_password": super_admin_password,
        },
    )
    if Path(file.filename or "").suffix.lower() != ".zip":
        raise HTTPException(status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE, detail="离线更新包必须是 .zip 文件")
    limit = max(20, min(int(settings.update_max_package_mb), 1024)) * 1024 * 1024
    content = await file.read(limit + 1)
    if not content or len(content) > limit:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail=f"离线更新包必须小于 {limit // 1024 // 1024} MB")
    incoming = update_run_root() / f"offline-{uuid4()}"
    incoming.mkdir(parents=True, exist_ok=True)
    package = incoming / "student-management-update.zip"
    package.write_bytes(content)
    return _schedule_update(db, user, approver, request, offline_package=package)


@app.get("/api/system/backups")
def list_backups(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> list[dict[str, Any]]:
    records = list(db.scalars(select(SystemBackup).order_by(SystemBackup.created_at.desc()).limit(100)))
    return [_serialize_backup(record) for record in records]


@app.post("/api/system/backups")
def create_backup(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    backup = create_database_backup(db, user)
    audit(
        db,
        "create_database_backup",
        "system_backup",
        backup.id,
        actor=user,
        after={"status": backup.status, "file_name": backup.file_name, "size_bytes": backup.size_bytes, "error": backup.error_message},
        request=request,
    )
    db.commit()
    if backup.status != "completed":
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=backup.error_message or "数据库备份失败")
    return _serialize_backup(backup)


@app.post("/api/system/backups/{backup_id}/validate")
def validate_system_backup(
    backup_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    backup = db.get(SystemBackup, backup_id)
    if not backup or backup.status != "completed":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="备份记录不存在")
    try:
        manifest = validate_backup(backup)
        backup.validation_status = "valid"
        backup.validated_at = utcnow()
        backup.manifest = manifest
        audit(db, "validate_database_backup", "system_backup", backup.id, actor=user, after={"valid": True, "storage_files": len(manifest.get("storage_files") or [])}, request=request)
        db.commit()
    except Exception as exc:
        backup.validation_status = "invalid"
        backup.error_message = str(exc)[:1000]
        db.commit()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"备份校验失败: {exc}") from exc
    return _serialize_backup(backup)


@app.post("/api/system/backups/{backup_id}/drill")
def run_backup_restore_drill(
    backup_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    backup = db.get(SystemBackup, backup_id)
    if not backup or backup.status != "completed":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="备份记录不存在")
    try:
        result = drill_restore_backup(backup)
        backup.validation_status = "valid"
        backup.validated_at = utcnow()
        audit(db, "drill_database_backup_restore", "system_backup", backup.id, actor=user, after=result, request=request)
        db.commit()
    except Exception as exc:
        backup.validation_status = "invalid"
        backup.error_message = str(exc)[:1000]
        db.commit()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"隔离恢复演练失败: {exc}") from exc
    return {"ok": True, "message": "隔离恢复演练通过，当前数据库没有被修改", "result": result}


@app.post("/api/system/backups/{backup_id}/restore")
def restore_system_backup(
    backup_id: int,
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    if str(payload.get("confirmation_phrase") or "").strip() != "恢复备份":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请输入“恢复备份”确认")
    backup = db.get(SystemBackup, backup_id)
    if not backup or backup.status != "completed":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="备份记录不存在")
    try:
        validate_backup(backup)
        recovery_point = create_database_backup(db, user)
        if recovery_point.status != "completed":
            raise RuntimeError(recovery_point.error_message or "恢复前回滚备份创建失败")
        db.commit()
        result = restore_backup(db, backup)
        audit(
            db,
            "restore_database_backup",
            "system_backup",
            backup.id,
            actor=user,
            after={"recovery_point": recovery_point.file_name, "validation": "passed", **result},
            request=request,
        )
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"恢复失败: {exc}") from exc
    return {"ok": True, "result": result, "message": "已恢复学生档案及其相关资料。账号、审计、AI 记录、系统设置和全部备份记录均已保留；恢复前状态已另存为回滚备份。"}


@app.delete("/api/system/backups/{backup_id}")
def delete_system_backup(
    backup_id: int,
    payload: dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    if str(payload.get("confirmation_phrase") or "").strip() != "删除备份":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请输入“删除备份”进行二次确认")
    backup = db.get(SystemBackup, backup_id)
    if not backup:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="备份记录不存在")
    snapshot = {
        "file_name": backup.file_name,
        "status": backup.status,
        "size_bytes": backup.size_bytes,
        "offsite_status": (backup.manifest or {}).get("offsite", {}).get("status") if isinstance((backup.manifest or {}).get("offsite"), dict) else None,
    }
    try:
        result = delete_database_backup(backup)
        audit(db, "delete_database_backup", "system_backup", backup.id, actor=user, before=snapshot, after=result, request=request)
        db.delete(backup)
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"删除备份失败: {exc}") from exc
    return {"ok": True, "message": "备份文件及记录已删除", **result}


@app.get("/api/system/backups/{backup_id}/download")
def download_backup(
    backup_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
):
    backup = db.get(SystemBackup, backup_id)
    if not backup or backup.status != "completed" or not backup.storage_path or not backup.file_name:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="备份文件不存在")
    target = Path(backup.storage_path).resolve()
    backup_root = settings.backup_path.resolve()
    if backup_root not in target.parents or not target.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="备份文件不存在")
    audit(db, "download_database_backup", "system_backup", backup.id, actor=user, after={"file_name": backup.file_name})
    db.commit()
    return FileResponse(target, filename=backup.file_name, media_type="application/octet-stream")


def _directory_usage(path: Path, suffixes: set[str] | None = None) -> dict[str, int]:
    if not path.exists():
        return {"files": 0, "size_bytes": 0}
    files = 0
    size_bytes = 0
    for item in path.rglob("*"):
        if not item.is_file() or (suffixes and item.suffix.lower() not in suffixes):
            continue
        try:
            size_bytes += item.stat().st_size
            files += 1
        except OSError:
            continue
    return {"files": files, "size_bytes": size_bytes}


@app.get("/api/system/info")
def system_information(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, Any]:
    database_dialect = engine.dialect.name
    disk = shutil.disk_usage(settings.export_path.resolve())
    return {
        "release": APP_RELEASE,
        "database": {
            "dialect": database_dialect,
            "tables": len(inspect(engine).get_table_names()),
            "students": int(db.scalar(select(func.count()).select_from(Student)) or 0),
        },
        "storage": {
            "exports": _directory_usage(settings.export_path, {".xlsx", ".csv"}),
            "originals": _directory_usage(settings.storage_path),
            "backups": _directory_usage(settings.backup_path),
            "disk_free_bytes": disk.free,
            "disk_total_bytes": disk.total,
        },
        "security": {
            "environment": settings.environment,
            "https_required": settings.is_production,
            "cookie_secure": settings.cookie_secure,
            "idle_logout_minutes": 5,
            "data_encryption_enabled": bool(settings.data_encryption_key),
        },
    }


@app.post("/api/system/maintenance/cleanup-exports")
def cleanup_old_exports(
    request: Request,
    payload: dict[str, Any] | None = Body(default=None),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, Any]:
    require_csrf(request)
    try:
        retention_days = int((payload or {}).get("retention_days", 30))
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="保留天数必须是数字") from exc
    if not 1 <= retention_days <= 3650:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="保留天数需在 1 至 3650 天之间")
    cutoff = china_now().timestamp() - timedelta(days=retention_days).total_seconds()
    deleted_files = 0
    freed_bytes = 0
    for item in settings.export_path.glob("*"):
        if not item.is_file() or item.suffix.lower() not in {".xlsx", ".csv"}:
            continue
        try:
            stat = item.stat()
            if stat.st_mtime >= cutoff:
                continue
            item.unlink()
            deleted_files += 1
            freed_bytes += stat.st_size
        except OSError:
            continue
    result = {"retention_days": retention_days, "deleted_files": deleted_files, "freed_bytes": freed_bytes}
    audit(db, "cleanup_export_files", "export_storage", "generated_exports", actor=user, after=result, request=request)
    db.commit()
    return result


def _audit_reversible(action: str, entity_type: str, before: dict[str, Any] | None, after: dict[str, Any] | None) -> bool:
    if action == "create" and entity_type == "student":
        return bool(after)
    if action in {"update", "restore_student_version", "delete_student"} and entity_type == "student":
        return bool(before and after)
    if action == "restore_deleted_student" and entity_type == "deleted_student":
        return bool(after)
    if action == "clear_all_students_high_risk" and entity_type == "student_archive":
        return bool(after and isinstance(after.get("recycle_ids"), list))
    if action in {"create_administrator", "create_import_template", "create_export_template"}:
        return bool(after)
    if action in {"update_administrator", "update_system_settings", "update_system_controls", "update_data_scope", "update_import_template", "update_export_template", "update_quality_issue", "update_source_document"}:
        return bool(before and after)
    if action in {"delete_import_template", "delete_export_template"}:
        return bool(before and after)
    if action == "save_student_filter" and entity_type == "saved_student_filter":
        return bool(after)
    if action == "delete_student_filter" and entity_type == "saved_student_filter":
        return bool(before)
    if action == "delete_related_info_card":
        return bool(before and after and {"student_id", "source_document_id", "import_batch_id", "imported_by_id", "title", "excel_payload"}.issubset(before))
    if action == "delete_source_related_cards":
        return bool(before and after and isinstance(before.get("cards"), list))
    return (action, entity_type) in {("import_excel", "import_batch"), ("commit_excel_import", "import_batch")} and bool(after)


def _student_snapshot_payload(snapshot: dict[str, Any]) -> dict[str, Any]:
    payload = {field: snapshot.get(field) for field in STUDENT_FIELDS if field in snapshot}
    for field in DATE_FILTER_FIELDS:
        if isinstance(payload.get(field), str) and payload[field]:
            payload[field] = date.fromisoformat(payload[field])
    return payload


def _student_snapshot_matches(student: Student, snapshot: dict[str, Any]) -> bool:
    current = student_to_dict(student)
    return all(current.get(field) == snapshot.get(field) for field in STUDENT_FIELDS if field in snapshot)


def _snapshot_fields_match(current: dict[str, Any], expected: dict[str, Any], fields: tuple[str, ...]) -> bool:
    return all(current.get(field) == expected.get(field) for field in fields)


def _restore_student_snapshot(db: Session, student: Student, snapshot: dict[str, Any], actor: User) -> dict[str, Any]:
    before = student_to_dict(student)
    payload = _student_snapshot_payload(snapshot)
    changed = [field for field in payload if field != "student_no" and before.get(field) != snapshot.get(field)]
    for field, value in payload.items():
        if field != "student_no":
            setattr(student, field, value)
    if changed:
        student.row_version += 1
        db.flush()
        record_student_version(db, student, actor, changed)
    return before


def _restore_data_scope(db: Session, account_id: int, snapshot: dict[str, Any] | list[dict[str, Any]], actor: User) -> None:
    state = snapshot if isinstance(snapshot, dict) else {"rules": snapshot}
    rules = state.get("rules") if isinstance(state.get("rules"), list) else []
    scope_mode = str(state.get("scope_mode") or "all")
    existing = db.scalar(select(UserDataScope).where(UserDataScope.user_id == account_id))
    if not rules:
        if scope_mode == "unconfigured":
            if existing:
                db.delete(existing)
        else:
            primary = {key: None for key in ("school", "college", "school_major", "current_class")}
            if existing:
                for key, value in primary.items():
                    setattr(existing, key, value)
                existing.updated_by_id = actor.id
            else:
                db.add(UserDataScope(user_id=account_id, updated_by_id=actor.id, **primary))
    else:
        primary = rules[0]
        if existing:
            for key in ("school", "college", "school_major", "current_class"):
                setattr(existing, key, primary.get(key))
            existing.updated_by_id = actor.id
        else:
            db.add(UserDataScope(user_id=account_id, updated_by_id=actor.id, **{key: primary.get(key) for key in ("school", "college", "school_major", "current_class")}))
    db.execute(delete(UserDataScopeRule).where(UserDataScopeRule.user_id == account_id))
    for rule in rules[1:]:
        db.add(UserDataScopeRule(user_id=account_id, created_by_id=actor.id, **{key: rule.get(key) for key in ("school", "college", "school_major", "current_class")}))
    db.flush()


def _undo_audit_record(db: Session, record: AuditLog, actor: User) -> dict[str, Any]:
    action = record.action
    before = record.before_data or {}
    after = record.after_data or {}

    if action in {"import_excel", "commit_excel_import"} and record.entity_type == "import_batch":
        batch = db.get(ImportBatch, int(record.entity_id))
        if not batch:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入批次不存在，无法撤回")
        result = rollback_import_batch(db, batch, actor)
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": result}

    if action == "clear_all_students_high_risk" and record.entity_type == "student_archive":
        restored_count = 0
        for recycle_id in after.get("recycle_ids") or []:
            recycle_record = db.get(DeletedStudent, int(recycle_id))
            if recycle_record and recycle_record.restored_at is None:
                restore_deleted_student(db, recycle_record, actor)
                restored_count += 1
        if restored_count != int(after.get("deleted_student_count") or 0):
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="部分学生回收站快照已被处理，无法安全批量撤回")
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": {"restored_students": restored_count}}

    if record.entity_type == "student":
        student = db.get(Student, int(record.entity_id))
        expected = after.get("student") if action == "restore_student_version" else after
        if action == "create":
            if not student or not _student_snapshot_matches(student, expected):
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="学生档案已经发生后续变化，无法直接撤回")
            recycle_record = permanently_delete_student(db, student, actor)
            return {"action": action, "entity_type": "student", "entity_id": record.entity_id, "result": {"moved_to_recycle_bin": recycle_record.id}}
        if action in {"update", "restore_student_version"}:
            if not student or not _student_snapshot_matches(student, expected):
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="学生档案已经发生后续变化，无法直接撤回")
            _restore_student_snapshot(db, student, before, actor)
            return {"action": action, "entity_type": "student", "entity_id": record.entity_id, "result": {"restored": True}}
        if action == "delete_student":
            recycle_id = int(after.get("recycle_id") or 0)
            recycle_record = db.get(DeletedStudent, recycle_id)
            if not recycle_record:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生回收站快照不存在，无法撤回")
            restored = restore_deleted_student(db, recycle_record, actor)
            return {"action": action, "entity_type": "student", "entity_id": record.entity_id, "result": {"restored_student_id": restored.id}}
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该学生操作暂不支持审计撤回")

    if action == "restore_deleted_student" and record.entity_type == "deleted_student":
        student_id = int(after.get("student_id") or 0)
        student = db.get(Student, student_id)
        if not student:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="恢复后的学生档案不存在")
        recycle_record = permanently_delete_student(db, student, actor)
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": {"moved_to_recycle_bin": recycle_record.id}}

    if action in {"update_system_settings", "update_administrator"} and record.entity_type == "user":
        account = db.get(User, int(record.entity_id))
        if not account or account.username != after.get("username") or account.display_name != after.get("display_name"):
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="账号已经发生后续变化，无法直接撤回")
        account.username = before.get("username") or account.username
        account.display_name = before.get("display_name") or account.display_name
        if action == "update_administrator":
            account.role = Role(before.get("role") or account.role.value)
            account.permissions = before.get("permissions")
            account.session_version = max(1, int(account.session_version or 1)) + 1
        db.flush()
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": {"restored": True, "password_unchanged": True}}

    if action == "create_administrator" and record.entity_type == "user":
        account = db.get(User, int(record.entity_id))
        if not account:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="账号不存在")
        if account.username != after.get("username") or account.display_name != after.get("display_name") or account.role.value != after.get("role") or (account.permissions or []) != (after.get("permissions") or []) or int(account.session_version or 1) != 1:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="账号已经发生后续变化，无法直接撤回")
        if db.scalar(select(UserDataScope).where(UserDataScope.user_id == account.id)):
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该账号已有数据范围配置，无法直接撤回")
        db.delete(account)
        db.flush()
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": {"deleted": True}}

    if action == "update_system_controls" and record.entity_type == "system_preference":
        if get_controls(db) != {key: bool(value) for key, value in after.items()}:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="系统控制项已经发生后续变化，无法直接撤回")
        controls = set_controls(db, before, actor)
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": controls}

    if action == "update_data_scope" and record.entity_type == "user":
        account_id = int(record.entity_id)
        account = db.get(User, account_id)
        if not account or get_user_scopes(db, account) != (after.get("rules") or []) or (after.get("scope_mode") and _scope_mode(db, account) != after.get("scope_mode")):
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="数据范围已经发生后续变化，无法直接撤回")
        _restore_data_scope(db, account_id, before, actor)
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": {"restored": True}}

    if record.entity_type == "saved_student_filter" and action in {"save_student_filter", "delete_student_filter"}:
        filter_id = int(record.entity_id)
        saved_filter = db.get(SavedStudentFilter, filter_id)
        if action == "save_student_filter":
            if not saved_filter or saved_filter.user_id != record.actor_id or not _snapshot_fields_match({"name": saved_filter.name, "filters": saved_filter.filters or {}}, after, ("name", "filters")):
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="常用筛选已经发生后续变化，无法直接撤回")
            if before is None:
                db.delete(saved_filter)
            else:
                saved_filter.name = before.get("name") or saved_filter.name
                saved_filter.filters = before.get("filters") or {}
        else:
            if saved_filter:
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="同一常用筛选已经重新存在，无法直接撤回")
            if not record.actor_id:
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="原常用筛选缺少所属账号，无法直接撤回")
            db.add(SavedStudentFilter(id=filter_id, user_id=record.actor_id, name=before.get("name") or "已恢复筛选", filters=before.get("filters") or {}))
        db.flush()
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": {"restored": True}}

    if action in {"create_import_template", "update_import_template", "delete_import_template"} and record.entity_type == "import_template":
        template_id = int(record.entity_id)
        template = db.get(ImportMappingTemplate, template_id)
        if action == "create_import_template":
            if not template:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入模板不存在")
            if not _snapshot_fields_match(_serialize_import_template(template), after, ("name", "mapping", "required_fields", "default_mode", "update_policy")):
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="导入模板已经发生后续变化，无法直接撤回")
            db.delete(template)
        elif action == "update_import_template":
            if not template or not _snapshot_fields_match(_serialize_import_template(template), after, ("name", "mapping", "required_fields", "default_mode", "update_policy")):
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="导入模板已经发生后续变化，无法直接撤回")
            template.name = before["name"]
            template.mapping = before.get("mapping") or {}
            template.required_fields = before.get("required_fields") or []
            template.default_mode = before.get("default_mode") or "upsert"
            template.update_policy = before.get("update_policy") or "overwrite"
        else:
            if template:
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="同一导入模板已经被重新创建")
            db.add(ImportMappingTemplate(id=template_id, name=before["name"], mapping=before.get("mapping") or {}, required_fields=before.get("required_fields") or [], default_mode=before.get("default_mode") or "upsert", update_policy=before.get("update_policy") or "overwrite", created_by_id=before.get("created_by_id")))
        db.flush()
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": {"restored": True}}

    if action in {"create_export_template", "update_export_template", "delete_export_template"} and record.entity_type == "export_template":
        template_id = int(record.entity_id)
        template = db.get(ExportTemplate, template_id)
        if action == "create_export_template":
            if not template:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导出模板不存在")
            if not _snapshot_fields_match(_serialize_export_template(template), after, ("name", "fields", "filters", "include_provenance", "mask_sensitive")):
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="导出模板已经发生后续变化，无法直接撤回")
            db.delete(template)
        elif action == "update_export_template":
            if not template or not _snapshot_fields_match(_serialize_export_template(template), after, ("name", "fields", "filters", "include_provenance", "mask_sensitive")):
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="导出模板已经发生后续变化，无法直接撤回")
            template.name = before["name"]
            template.fields = before.get("fields") or []
            template.filters = before.get("filters") or {}
            template.include_provenance = bool(before.get("include_provenance", True))
            template.mask_sensitive = bool(before.get("mask_sensitive", False))
        else:
            if template:
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="同一导出模板已经被重新创建")
            db.add(ExportTemplate(id=template_id, name=before["name"], fields=before.get("fields") or [], filters=before.get("filters") or {}, include_provenance=bool(before.get("include_provenance", True)), mask_sensitive=bool(before.get("mask_sensitive", False)), created_by_id=before.get("created_by_id")))
        db.flush()
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": {"restored": True}}

    if action == "update_source_document" and record.entity_type == "source_document":
        document = db.get(SourceDocument, int(record.entity_id))
        if not document or document.status == "deleted":
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="原始资料不存在或已经删除")
        if list(document.tags or []) != list(after.get("tags") or []) or document.status != after.get("status"):
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="原始资料已经发生后续变化，无法直接撤回")
        document.tags = list(before.get("tags") or [])
        document.status = before.get("status") or "active"
        document.archived_at = None if document.status == "active" else document.archived_at
        db.flush()
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": {"restored": True}}

    if action == "delete_related_info_card" and record.entity_type == "student_related_info_card":
        if db.get(StudentRelatedInfoCard, int(record.entity_id)):
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="相关信息词条已经重新存在")
        student_id = int(before.get("student_id") or 0)
        if not db.get(Student, student_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="关联学生不存在")
        imported_at = before.get("imported_at")
        if isinstance(imported_at, str) and imported_at:
            imported_at = datetime.fromisoformat(imported_at)
        db.add(StudentRelatedInfoCard(id=int(record.entity_id), student_id=student_id, source_document_id=int(before["source_document_id"]), import_batch_id=int(before["import_batch_id"]), imported_by_id=int(before["imported_by_id"]), title=before.get("title") or "相关信息", excel_payload=before.get("excel_payload") or {}, imported_at=imported_at or utcnow()))
        db.flush()
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": {"restored": True}}

    if action == "delete_source_related_cards" and record.entity_type == "source_document":
        cards = before.get("cards") or []
        for item in cards:
            if not db.get(StudentRelatedInfoCard, int(item.get("id") or 0)):
                imported_at = item.get("imported_at")
                if isinstance(imported_at, str) and imported_at:
                    imported_at = datetime.fromisoformat(imported_at)
                db.add(StudentRelatedInfoCard(id=int(item["id"]), student_id=int(item["student_id"]), source_document_id=int(item["source_document_id"]), import_batch_id=int(item["import_batch_id"]), imported_by_id=int(item["imported_by_id"]), title=item.get("title") or "相关信息", excel_payload=item.get("excel_payload") or {}, imported_at=imported_at or utcnow()))
        db.flush()
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": {"restored_cards": len(cards)}}

    if action == "update_quality_issue" and record.entity_type == "quality_issue_case":
        case = db.get(QualityIssueCase, int(record.entity_id))
        if not case:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="质量问题不存在")
        if case.status != after.get("status") or case.assignee_id != after.get("assignee_id") or case.resolution_note != after.get("resolution_note"):
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="质量问题已经发生后续变化，无法直接撤回")
        case.status = before.get("status") or "open"
        case.assignee_id = before.get("assignee_id")
        case.resolution_note = before.get("resolution_note")
        case.resolved_at = None if case.status == "open" else case.resolved_at
        db.flush()
        return {"action": action, "entity_type": record.entity_type, "entity_id": record.entity_id, "result": {"restored": True}}

    raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该审计记录缺少可安全恢复的完整快照，不能直接撤回")


@app.post("/api/audit/{audit_id}/undo")
def undo_audit_change(
    audit_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    _require_capability(user, "audit_view", "审计查看")
    require_csrf(request)
    record = db.get(AuditLog, audit_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="审计记录不存在")
    if not _admin_can_view_teacher_record(db, user, record.actor_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="审计记录不存在")
    if db.scalar(select(AuditReversal).where(AuditReversal.audit_log_id == record.id)):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该改动已经撤回过")
    if not _audit_reversible(record.action, record.entity_type, record.before_data, record.after_data):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="该审计记录没有可安全恢复的完整快照")
    try:
        result = _undo_audit_record(db, record, user)
        reversal = AuditReversal(audit_log_id=record.id, undone_by_id=user.id, result=result)
        db.add(reversal)
        db.flush()
        audit(db, "undo_audit_change", "audit_log", record.id, actor=user, before={"action": record.action, "entity_type": record.entity_type, "entity_id": record.entity_id}, after={"reversal_id": reversal.id, "result": result}, request=request)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"ok": True, "audit_id": record.id, "result": result}


@app.get("/api/audit")
def audit_logs(
    action: str | None = None,
    actor_id: int | None = None,
    entity_type: str | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    limit: int = 200,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> list[dict[str, Any]]:
    _require_capability(user, "audit_view", "审计查看")
    statement = select(AuditLog)
    if action:
        statement = statement.where(AuditLog.action == action)
    if actor_id:
        statement = statement.where(AuditLog.actor_id == actor_id)
    if entity_type:
        statement = statement.where(AuditLog.entity_type == entity_type)
    if user.role == Role.ADMIN:
        statement = statement.where(AuditLog.actor_id.in_(select(User.id).where(User.role == Role.TEACHER)))
    if date_from:
        statement = statement.where(AuditLog.created_at >= date_from)
    if date_to:
        statement = statement.where(AuditLog.created_at <= date_to)
    records = list(db.scalars(statement.order_by(AuditLog.created_at.desc()).limit(min(max(limit, 1), 500))))
    actor_ids = {record.actor_id for record in records if record.actor_id}
    actors = {actor.id: actor for actor in db.scalars(select(User).where(User.id.in_(actor_ids)))} if actor_ids else {}
    record_ids = [record.id for record in records]
    reversals = {item.audit_log_id: item for item in db.scalars(select(AuditReversal).where(AuditReversal.audit_log_id.in_(record_ids)))} if record_ids else {}
    return [
        {
            "id": record.id,
            "action": record.action,
            "entity_type": record.entity_type,
            "entity_id": record.entity_id,
            "actor_id": record.actor_id,
            "actor": (actors[record.actor_id].display_name or actors[record.actor_id].username) if record.actor_id in actors else "系统",
            "created_at": record.created_at,
            "before": record.before_data,
            "after": record.after_data,
            "entry_hash": record.entry_hash,
            "can_undo": _audit_reversible(record.action, record.entity_type, record.before_data, record.after_data) and record.id not in reversals,
            "undone": record.id in reversals,
            "undone_at": reversals[record.id].created_at if record.id in reversals else None,
        }
        for record in records
    ]


@app.get("/api/auth/login-security-events")
def login_security_events(
    limit: int = 100,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> list[dict[str, Any]]:
    _require_capability(user, "audit_view", "审计查看")
    if user.role == Role.ADMIN:
        teacher_ids = select(User.id).where(User.role == Role.TEACHER)
    else:
        teacher_ids = None
    statement = select(LoginSecurityEvent)
    if teacher_ids is not None:
        statement = statement.where(LoginSecurityEvent.user_id.in_(teacher_ids))
    records = list(
        db.scalars(
            statement
            .order_by(LoginSecurityEvent.created_at.desc(), LoginSecurityEvent.id.desc())
            .limit(min(max(limit, 1), 300))
        )
    )
    user_ids = {record.user_id for record in records if record.user_id}
    users = {item.id: item for item in db.scalars(select(User).where(User.id.in_(user_ids)))} if user_ids else {}
    return [
        {
            "id": record.id,
            "username": record.username,
            "display_name": (users[record.user_id].display_name or users[record.user_id].username) if record.user_id in users else record.username,
            "event_type": record.event_type,
            "ip_address": record.ip_address,
            "network_key": record.network_key,
            "device_label": record.device_label,
            "is_unusual": record.is_unusual,
            "created_at": record.created_at,
        }
        for record in records
    ]


@app.get("/api/audit/verify")
def audit_verify(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
) -> dict[str, Any]:
    _require_capability(user, "audit_view", "审计查看")
    return verify_audit_chain(db)


@app.get("/api/audit/export")
def export_audit_logs(
    action: str | None = None,
    actor_id: int | None = None,
    entity_type: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN, Role.ADMIN)),
):
    _require_capability(user, "audit_view", "审计查看")
    statement = select(AuditLog)
    if action:
        statement = statement.where(AuditLog.action == action)
    if actor_id:
        statement = statement.where(AuditLog.actor_id == actor_id)
    if entity_type:
        statement = statement.where(AuditLog.entity_type == entity_type)
    if user.role == Role.ADMIN:
        statement = statement.where(AuditLog.actor_id.in_(select(User.id).where(User.role == Role.TEACHER)))
    records = list(db.scalars(statement.order_by(AuditLog.created_at.desc()).limit(10000)))
    actor_ids = {item.actor_id for item in records if item.actor_id}
    actors = {item.id: item for item in db.scalars(select(User).where(User.id.in_(actor_ids)))} if actor_ids else {}
    content = io.StringIO()
    writer = csv.writer(content)
    writer.writerow(["时间", "操作人", "动作", "对象类型", "对象 ID", "变更前", "变更后", "校验哈希"])
    for record in records:
        actor = actors.get(record.actor_id)
        writer.writerow([record.created_at.isoformat(), (actor.display_name or actor.username) if actor else "系统", record.action, record.entity_type, record.entity_id, json.dumps(record.before_data, ensure_ascii=False, default=str), json.dumps(record.after_data, ensure_ascii=False, default=str), record.entry_hash or ""])
    audit(db, "export_audit_logs", "audit_log", "filtered", actor=user, after={"count": len(records), "action": action, "actor_id": actor_id, "entity_type": entity_type})
    db.commit()
    return Response(content.getvalue().encode("utf-8-sig"), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=audit_logs.csv"})


@app.post("/api/audit/retention")
def apply_audit_retention(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.SUPER_ADMIN)),
) -> dict[str, int]:
    require_csrf(request)
    cutoff = utcnow() - timedelta(days=max(30, settings.audit_retention_days))
    count = db.execute(delete(AuditLog).where(AuditLog.created_at < cutoff)).rowcount or 0
    audit(db, "apply_audit_retention", "audit_log", "retention", actor=user, after={"removed": count, "cutoff": cutoff.isoformat()}, request=request)
    db.commit()
    return {"removed": count}

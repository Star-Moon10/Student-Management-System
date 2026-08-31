from pathlib import Path
import re
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.time import as_china_time, china_now
from app.models import FieldProvenance, ImportBatch, SourceDocument, Student
from app.services.students import build_student_query


def safe_export_filename_stem(value: str | None) -> str:
    """Keep a human-friendly export title while making it safe on Windows."""
    stem = str(value or "").strip()
    if stem.lower().endswith(".xlsx"):
        stem = stem[:-5]
    stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", stem)
    stem = re.sub(r"\s+", " ", stem).strip(" .")
    return stem[:60] or "students"

EXPORT_FIELD_HEADERS = {
    "student_no": "学号",
    "candidate_no": "考生号",
    "full_name": "姓名",
    "gender": "性别",
    "national_id": "身份证号",
    "date_of_birth": "出生日期",
    "student_origin": "生源地",
    "ethnicity": "民族",
    "political_status": "政治面貌",
    "enrollment_date": "入学日期",
    "graduation_year": "毕业年份",
    "graduation_date": "毕业日期",
    "urban_rural_origin": "城乡生源",
    "pre_enrollment_archive_unit": "入学前档案所在单位",
    "archive_transferred": "档案是否转入学校",
    "pre_enrollment_police_station": "入学前户口所在地派出所",
    "household_registration_transferred": "户口是否转入学校",
    "education_level": "学历层次",
    "program_duration": "学制",
    "school": "所属学校",
    "college": "所属学院",
    "school_major": "学校专业",
    "major_direction": "专业方向",
    "current_class": "所在班级",
    "training_mode": "培养方式",
    "commissioned_unit": "委培单位",
    "hardship_category": "困难生类别",
    "normal_student_category": "师范生类别",
    "mobile_phone": "手机号码",
    "electronic_email": "电子邮箱",
    "qq_number": "QQ号码",
    "family_phone": "家庭电话",
    "family_postcode": "家庭邮编",
    "family_address": "家庭地址",
    "poverty_county_52": "是否52个贫困县",
    "poverty_county_province": "贫困县所在省",
    "poverty_county_city": "贫困县所在市",
    "poverty_county_district": "贫困县所在县",
    "registered_poor": "是否建档立卡",
    "study_mode": "学习形式",
    "vocational_expansion_flag": "高职扩招考生标志",
    "remarks": "备注",
}
FULL_EXPORT_FIELDS = tuple(EXPORT_FIELD_HEADERS)
SENSITIVE_EXPORT_FIELDS = {"national_id", "mobile_phone", "electronic_email", "qq_number", "family_phone", "family_postcode", "family_address"}
FIELD_WIDTHS = {
    "student_no": 18,
    "candidate_no": 18,
    "full_name": 16,
    "gender": 10,
    "national_id": 22,
    "date_of_birth": 14,
    "student_origin": 24,
    "ethnicity": 12,
    "political_status": 16,
    "enrollment_date": 14,
    "graduation_year": 14,
    "graduation_date": 14,
    "urban_rural_origin": 14,
    "pre_enrollment_archive_unit": 30,
    "archive_transferred": 18,
    "pre_enrollment_police_station": 30,
    "household_registration_transferred": 18,
    "education_level": 16,
    "program_duration": 12,
    "school": 22,
    "college": 22,
    "school_major": 28,
    "major_direction": 22,
    "current_class": 16,
    "training_mode": 16,
    "commissioned_unit": 28,
    "hardship_category": 18,
    "normal_student_category": 18,
    "mobile_phone": 18,
    "electronic_email": 28,
    "qq_number": 16,
    "family_phone": 18,
    "family_postcode": 14,
    "family_address": 34,
    "poverty_county_52": 20,
    "poverty_county_province": 18,
    "poverty_county_city": 18,
    "poverty_county_district": 18,
    "registered_poor": 16,
    "study_mode": 16,
    "vocational_expansion_flag": 22,
    "remarks": 42,
}


def _selected_export_fields(fields: list[str] | None) -> list[str]:
    if not fields:
        return list(FULL_EXPORT_FIELDS)
    selected = [field for field in fields if field in EXPORT_FIELD_HEADERS and field not in {"student_no", "full_name"}]
    return ["student_no", "full_name", *dict.fromkeys(selected)]


def _export_value(student: Student, field: str):
    value = getattr(student, field)
    if field in {"date_of_birth", "enrollment_date", "graduation_date"} and value:
        return value.isoformat()
    return value


def _masked_export_value(value):
    if value in (None, ""):
        return value
    text = str(value)
    if len(text) <= 5:
        return "*" * len(text)
    return f"{text[:3]}{'*' * max(2, len(text) - 5)}{text[-2:]}"


def create_student_export(
    db: Session,
    keyword: str | None = None,
    current_class: str | None = None,
    school_major: str | None = None,
    college: str | None = None,
    school: str | None = None,
    filters: dict[str, str] | None = None,
    student_ids: list[int] | None = None,
    include_provenance: bool = True,
    fields: list[str] | None = None,
    scope: dict[str, str] | list[dict[str, str]] | None = None,
    mask_sensitive: bool = False,
    filename_stem: str | None = None,
) -> Path:
    student_query = build_student_query(keyword, current_class, school_major, college, school, filters, scope=scope)
    if student_ids:
        student_query = student_query.where(Student.id.in_(student_ids))
    students = list(db.scalars(student_query))
    selected_fields = _selected_export_fields(fields)
    workbook = Workbook()
    # Some Excel protected-view installations reject openpyxl's empty
    # workbookProtection element even though the workbook is not protected.
    workbook.security = None
    worksheet = workbook.active
    worksheet.title = "学生汇总"
    worksheet.append([EXPORT_FIELD_HEADERS[field] for field in selected_fields])
    header_fill = PatternFill("solid", fgColor="0F766E")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for student in students:
        worksheet.append([
            _masked_export_value(_export_value(student, field)) if mask_sensitive and field in SENSITIVE_EXPORT_FIELDS else _export_value(student, field)
            for field in selected_fields
        ])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, field in enumerate(selected_fields, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = FIELD_WIDTHS[field]

    if include_provenance:
        source_sheet = workbook.create_sheet("数据来源")
        source_sheet.append(["学号", "姓名", "字段", "原始文件", "工作表", "行", "列", "单元格/位置", "原始值", "记录时间"])
        for cell in source_sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
        student_ids = [student.id for student in students]
        if student_ids:
            provenance_rows = db.execute(
                select(FieldProvenance, Student, SourceDocument.original_filename)
                .join(Student, FieldProvenance.student_id == Student.id)
                .outerjoin(SourceDocument, FieldProvenance.source_document_id == SourceDocument.id)
                .where(FieldProvenance.student_id.in_(student_ids))
                .where(FieldProvenance.field_name.in_(selected_fields))
                .order_by(FieldProvenance.recorded_at.desc())
            )
            for provenance, student, source_filename in provenance_rows:
                source_sheet.append(
                    [
                        student.student_no,
                        student.full_name,
                        provenance.field_name,
                        source_filename or "平台编辑",
                        provenance.source_sheet,
                        provenance.source_row,
                        provenance.source_column,
                        provenance.source_locator,
                        provenance.raw_value,
                        as_china_time(provenance.recorded_at).strftime("%Y-%m-%d %H:%M:%S") if provenance.recorded_at else "",
                    ]
                )
        source_sheet.freeze_panes = "A2"
        source_sheet.auto_filter.ref = source_sheet.dimensions
        for column, width in {"A": 18, "B": 16, "C": 18, "D": 18, "E": 18, "F": 10, "G": 10, "H": 24, "I": 36, "J": 22}.items():
            source_sheet.column_dimensions[column].width = width

    settings = get_settings()
    settings.export_path.mkdir(parents=True, exist_ok=True)
    filename = f"{safe_export_filename_stem(filename_stem)}_{china_now():%Y%m%d_%H%M%S}.xlsx"
    target = settings.export_path / filename
    if target.exists():
        target = settings.export_path / f"{target.stem}_{uuid4().hex[:8]}{target.suffix}"
    workbook.save(target)
    return target


def create_import_batch_report(batch: ImportBatch, filename: str | None = None) -> Path:
    """Create a portable XLSX report for a completed import batch."""
    workbook = Workbook()
    workbook.security = None
    summary = workbook.active
    summary.title = "导入报告"
    header_fill = PatternFill("solid", fgColor="0F766E")
    summary.append(["项目", "内容"])
    for cell in summary[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    values = [
        ("批次 ID", batch.id),
        ("原始文件", filename or "已删除文件"),
        ("导入方式", batch.mode),
        ("导入状态", batch.status.value if hasattr(batch.status, "value") else batch.status),
        ("总行数", batch.total_rows),
        ("新增", batch.created_rows),
        ("更新", batch.updated_rows),
        ("跳过", batch.skipped_rows),
        ("错误", batch.error_rows),
        ("撤销状态", batch.rollback_status or "不支持"),
        ("完成时间", batch.completed_at.isoformat() if batch.completed_at else "-"),
    ]
    for key, value in values:
        summary.append([key, value])
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 60

    errors = workbook.create_sheet("错误明细")
    errors.append(["行号", "原因"])
    for cell in errors[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for item in batch.errors or []:
        errors.append([item.get("row"), item.get("message")])
    errors.column_dimensions["A"].width = 12
    errors.column_dimensions["B"].width = 70

    rollback = workbook.create_sheet("可撤销变更")
    rollback.append(["类型", "学生学号", "字段", "状态"])
    for cell in rollback[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    data = batch.rollback_data or {}
    for item in data.get("created", []):
        rollback.append(["新增学生", item.get("student_no"), "全部档案", batch.rollback_status or "-"])
    for item in data.get("updated", []):
        rollback.append(["更新学生", item.get("student_no"), "、".join(item.get("changed_fields") or []), batch.rollback_status or "-"])
    for column, width in {"A": 18, "B": 20, "C": 50, "D": 18}.items():
        rollback.column_dimensions[column].width = width

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    settings = get_settings()
    settings.export_path.mkdir(parents=True, exist_ok=True)
    target = settings.export_path / f"import_report_{batch.id}_{china_now():%Y%m%d_%H%M%S}_{uuid4().hex[:8]}.xlsx"
    workbook.save(target)
    return target
